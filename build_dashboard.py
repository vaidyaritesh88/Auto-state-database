#!/usr/bin/env python3
"""
Janchor Auto Tracker - Dashboard Generator v2

Usage:
  python build_dashboard.py                    # Build using existing data.json
  python build_dashboard.py path/to/file.xlsx  # Convert Excel to data.json, then build
                                               # Supports both old format and Kotak format
"""
import json, os, sys, re

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_FILE = os.path.join(SCRIPT_DIR, 'data.json')
OUTPUT_FILE = os.path.join(SCRIPT_DIR, 'dashboard.html')

# ── OEM name cleanup (Kotak full legal names → short display names) ──────────
OEM_NAME_MAP = {
    'Maruti Suzuki India Ltd':'Maruti Suzuki','Hyundai Motor India Ltd':'Hyundai',
    'Tata Motors Ltd':'Tata Motors','Mahindra & Mahindra Ltd':'M&M',
    'Toyota Kirloskar Motor Pvt Ltd':'Toyota','Kia Motors':'Kia',
    'Honda Cars India Ltd':'Honda','SkodaAuto India Pvt Ltd':'Skoda',
    'Volkswagen India Pvt Ltd':'Volkswagen','Renault India Pvt Ltd':'Renault',
    'Nissan Motor India Pvt Ltd':'Nissan','Ford India Pvt Ltd':'Ford',
    'Fiat India Automobiles Pvt.Ltd':'Fiat India',
    'General Motors India Pvt Ltd':'General Motors',
    'Hindustan Motor Finance Corporation Ltd':'Hindustan Motor',
    'Force Motors Ltd':'Force Motors','Isuzu Motors India Pvt Ltd':'Isuzu',
    'PCA Motors Pvt Ltd':'PCA Motors','MG Motor':'MG Motor',
    'Hero MotoCorp Ltd':'Hero',
    'Honda Motorcycle & Scooter India (Pvt) Ltd':'Honda',
    'TVS Motor Company Ltd':'TVS Motor','Bajaj Auto Ltd':'Bajaj',
    'Royal Enfield (A Unit of Eicher Motors Ltd)':'Royal Enfield',
    'Suzuki Motorcycle India Pvt Ltd':'Suzuki',
    'India Yamaha Motor Pvt Ltd':'Yamaha',
    'Ather Energy Pvt. Ltd':'Ather','Okinawa Autotech Pvt. Ltd':'Okinawa',
    'India Kawasaki Motors Pvt Ltd':'Kawasaki',
    'H-D Motor Company India Pvt Ltd':'Harley Davidson',
    'Mahindra Two Wheelers Ltd':'M&M',
    'Piaggio Vehicles Pvt Ltd':'Piaggio',
    'UM Lohia Two Wheelers Pvt Ltd':'UM Lohia',
    'Ashok Leyland Ltd':'Ashok Leyland','VECV- Eicher':'VECV',
    'Volvo group':'Volvo Group','SML Isuzu Ltd':'SML Isuzu',
    'TI Clean Mobility Pvt Ltd':'TI Clean Mobility',
    'Pinnacle Mobility Solutions Pvt Ltd':'Pinnacle Mobility',
    'Atul Auto Ltd':'Atul Auto',
}

def clean_oem(name):
    if name in OEM_NAME_MAP: return OEM_NAME_MAP[name]
    c = re.sub(r'\s*\(Pvt\)\s*', ' ', name)
    c = re.sub(r'\s*(Pvt\.?|Private)\s*(Ltd\.?|Limited)\s*$', '', c)
    c = re.sub(r'\s*(Ltd\.?|Limited)\s*$', '', c)
    c = re.sub(r'\s+India\s*$', '', c)
    return c.strip() or name

def to_num(v):
    if isinstance(v, (int, float)): return v
    try: return float(v)
    except: return 0

# ── Excel → data.json converter ──────────────────────────────────────────────
def convert_excel_to_json(excel_path):
    """Parse Excel (auto-detect old or Kotak format) and write data.json"""
    import openpyxl
    print(f"Reading Excel: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    sheets = wb.sheetnames
    print(f"  Sheets: {sheets}")

    # Detect format
    old_sheets = ['PVs - Raw data','2Ws - Raw data','3Ws - Raw data','M&HCVs - Raw data','LCVs - Raw data']
    if any(s in sheets for s in old_sheets):
        result = _parse_old_format(wb, old_sheets)
    else:
        result = _parse_kotak_format(wb)

    wb.close()
    if not result:
        print("ERROR: No data parsed!")
        sys.exit(1)
    with open(DATA_FILE, 'w') as f:
        json.dump(result, f)
    print(f"  Wrote {DATA_FILE}: {len(result['rows'])} rows, {len(result['quarters'])} quarters "
          f"({result['quarters'][0]} to {result['quarters'][-1]})")
    return result

def _parse_old_format(wb, old_sheets):
    """Parse old format with explicit Zone/State/Manufacturer columns"""
    SMAP = {'PVs - Raw data':'PV','2Ws - Raw data':'2W','3Ws - Raw data':'3W',
            'M&HCVs - Raw data':'MHCV','LCVs - Raw data':'LCV'}
    all_quarters, all_rows = None, []
    for sname, seg in SMAP.items():
        if sname not in wb.sheetnames: continue
        ws = wb[sname]
        rows = list(ws.iter_rows(min_row=1, values_only=True))
        # Find header with Q1FY17 pattern
        hdr_idx = -1
        for ri, row in enumerate(rows[:5]):
            for cell in row:
                if cell and re.match(r'^Q\dFY\d{2}$', str(cell)):
                    hdr_idx = ri; break
            if hdr_idx >= 0: break
        if hdr_idx < 0: continue
        headers = rows[hdr_idx]
        col_z = col_s = col_m = col_sub = -1; q_start = -1; quarters = []
        for ci, h in enumerate(headers):
            hs = str(h or '').strip(); hl = hs.lower()
            if hl == 'zone': col_z = ci
            elif hl == 'state': col_s = ci
            elif hl in ('manufacturer','oem'): col_m = ci
            elif hl in ('sub-segment','subsegment','sub_segment','sub segment'): col_sub = ci
            elif re.match(r'^Q\dFY\d{2}$', hs):
                if q_start < 0: q_start = ci
                quarters.append(hs)
        if col_z < 0 or col_s < 0 or col_m < 0 or q_start < 0: continue
        if all_quarters is None: all_quarters = quarters
        elif len(quarters) > len(all_quarters): all_quarters = quarters
        for row in rows[hdr_idx+1:]:
            zone = str(row[col_z] or '').strip()
            state = str(row[col_s] or '').strip()
            mfr = str(row[col_m] or '').strip()
            subseg = str(row[col_sub] or '').strip() if col_sub >= 0 else 'All'
            if not zone or not state or not mfr: continue
            if zone.lower() == 'zone' or state.lower() == 'state': continue
            vols = [to_num(row[q_start+q]) if q_start+q < len(row) else 0 for q in range(len(quarters))]
            if any(v > 0 for v in vols):
                all_rows.append([seg, subseg, zone, state, mfr] + vols)
    if not all_rows: return None
    return {'quarters': all_quarters,
            'columns': ['segment','subsegment','zone','state','manufacturer'] + ['vol_'+q.lower() for q in all_quarters],
            'rows': all_rows}

def _parse_kotak_format(wb):
    """Parse Kotak hierarchical format (Zone > State > OEM)"""
    ZONE_NAMES = ['North Zone','East Zone','West Zone','South Zone']
    ZONE_SHORT = {'North Zone':'North','East Zone':'East','West Zone':'West','South Zone':'South'}
    # Determine which sheets to process
    sheet_defs = []
    has_c = 'Cars' in wb.sheetnames; has_u = 'UVs' in wb.sheetnames
    has_m = 'Motorcycle' in wb.sheetnames; has_s = 'Scooters' in wb.sheetnames
    if has_c: sheet_defs.append(('Cars','PV','Cars'))
    if has_u: sheet_defs.append(('UVs','PV','UVs'))
    if not has_c and not has_u and 'PVs' in wb.sheetnames: sheet_defs.append(('PVs','PV','All'))
    if has_m: sheet_defs.append(('Motorcycle','2W','Motorcycle'))
    if has_s: sheet_defs.append(('Scooters','2W','Scooters'))
    if not has_m and not has_s and '2W' in wb.sheetnames: sheet_defs.append(('2W','2W','All'))
    if 'MHCVs' in wb.sheetnames: sheet_defs.append(('MHCVs','MHCV','All'))
    if 'LCVs' in wb.sheetnames: sheet_defs.append(('LCVs','LCV','All'))
    if '3W' in wb.sheetnames: sheet_defs.append(('3W','3W','All'))

    all_quarters, all_rows = None, []
    for sname, seg, sub in sheet_defs:
        ws = wb[sname]
        rows = list(ws.iter_rows(min_row=1, values_only=True))
        # Find header row with 1QFY16 pattern
        hdr_idx = -1
        for ri, row in enumerate(rows[:10]):
            for cell in row:
                if cell and re.match(r'^\dQFY\d{2}$', str(cell)):
                    hdr_idx = ri; break
            if hdr_idx >= 0: break
        if hdr_idx < 0: continue
        headers = rows[hdr_idx]
        name_col = -1; q_start = -1; quarters = []
        for ci, h in enumerate(headers):
            hs = str(h or '').strip()
            if re.match(r'^\dQFY\d{2}$', hs):
                if q_start < 0: q_start = ci
                quarters.append('Q' + hs[0] + hs[2:])  # 1QFY16 -> Q1FY16
        # Detect name column by finding a zone name
        for ri in range(hdr_idx+1, min(hdr_idx+15, len(rows))):
            row = rows[ri]
            if not row: continue
            for ci in range(min(5, len(row))):
                tv = str(row[ci] or '').strip()
                if tv in ZONE_NAMES: name_col = ci; break
            if name_col >= 0: break
        if name_col < 0 or q_start < 0: continue
        if all_quarters is None: all_quarters = quarters
        elif len(quarters) > len(all_quarters): all_quarters = quarters
        # Detect OEM list (first-repeat algorithm)
        first_zone = -1
        for ri in range(hdr_idx+1, len(rows)):
            n = str(rows[ri][name_col] or '').strip()
            if n in ZONE_NAMES: first_zone = ri; break
        if first_zone < 0: continue
        oem_list, seen = [], {}
        for ri in range(first_zone+1, len(rows)):
            n = str(rows[ri][name_col] or '').strip()
            if not n or n == '0': continue
            if n in seen: break
            seen[n] = True; oem_list.append(n)
        if len(oem_list) < 2: continue
        oem_list.pop()  # last = first state
        oem_set = set(oem_list)
        # Parse data
        cur_zone = ''; cur_state = ''; in_zone_block = False
        for ri in range(first_zone, len(rows)):
            row = rows[ri]; name = str(row[name_col] or '').strip()
            if not name or name == '0': continue
            nl = name.lower()
            if nl in ('total','grand total','all india','india'): break
            if name in ZONE_NAMES:
                cur_zone = ZONE_SHORT.get(name, name); cur_state = ''; in_zone_block = True; continue
            if name in oem_set:
                if in_zone_block or not cur_zone or not cur_state: continue
                vols = [to_num(row[q_start+q]) if q_start+q < len(row) else 0 for q in range(len(quarters))]
                if any(v > 0 for v in vols):
                    all_rows.append([seg, sub, cur_zone, cur_state, clean_oem(name)] + vols)
                continue
            cur_state = name; in_zone_block = False
        print(f"  {sname}: {seg}/{sub} parsed")
    if not all_rows: return None
    return {'quarters': all_quarters,
            'columns': ['segment','subsegment','zone','state','manufacturer'] + ['vol_'+q.lower() for q in all_quarters],
            'rows': all_rows}

# ── Main: optionally convert Excel, then build dashboard ─────────────────────
if len(sys.argv) > 1:
    convert_excel_to_json(sys.argv[1])

with open(DATA_FILE, 'r') as f:
    data_json = f.read()

html = '''<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Janchor Auto Tracker</title>
<script src="https://cdn.plot.ly/plotly-2.27.0.min.js"></script>
<script src="https://cdn.sheetjs.com/xlsx-0.20.1/package/dist/xlsx.full.min.js"></script>
<style>
*{margin:0;padding:0;box-sizing:border-box}
body{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,sans-serif;background:#f0f2f5;color:#1f2937;font-size:14px}
.header{background:linear-gradient(135deg,#1e3a5f 0%,#2563eb 100%);color:#fff;padding:12px 24px;display:flex;align-items:center;justify-content:space-between;flex-wrap:wrap;gap:8px}
.header h1{font-size:18px;font-weight:700;letter-spacing:0.5px;white-space:nowrap}
.seg-tabs{display:flex;gap:4px}
.seg-tab{padding:6px 16px;border-radius:20px;cursor:pointer;font-weight:600;font-size:13px;border:2px solid rgba(255,255,255,0.3);background:transparent;color:rgba(255,255,255,0.8);transition:all 0.2s}
.seg-tab:hover{background:rgba(255,255,255,0.15)}
.seg-tab.active{background:#fff;color:#1e3a5f;border-color:#fff}
.nav{background:#fff;border-bottom:1px solid #e5e7eb;padding:0 24px;display:flex;align-items:center;gap:24px}
.nav-tab{padding:12px 4px;cursor:pointer;font-weight:500;color:#6b7280;border-bottom:2px solid transparent;transition:all 0.2s;font-size:14px}
.nav-tab:hover{color:#1e3a5f}
.nav-tab.active{color:#2563eb;border-bottom-color:#2563eb;font-weight:600}
.main{max-width:1440px;margin:0 auto;padding:16px 24px}
.filter-bar{display:flex;gap:12px;align-items:center;flex-wrap:wrap;margin-bottom:16px;padding:12px 16px;background:#fff;border-radius:8px;box-shadow:0 1px 3px rgba(0,0,0,0.08)}
.filter-bar label{font-weight:600;font-size:12px;color:#6b7280;text-transform:uppercase;letter-spacing:0.5px}
.filter-bar select{padding:6px 28px 6px 10px;border:1px solid #d1d5db;border-radius:6px;font-size:13px;background:#fff;cursor:pointer;appearance:none;background-image:url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' width='12' height='12' fill='%236b7280' viewBox='0 0 16 16'%3E%3Cpath d='M8 11L3 6h10z'/%3E%3C/svg%3E");background-repeat:no-repeat;background-position:right 8px center}
.filter-bar select:focus{outline:none;border-color:#2563eb;box-shadow:0 0 0 2px rgba(37,99,235,0.2)}
.sep{width:1px;height:24px;background:#e5e7eb;margin:0 4px}
.kpi-row{display:grid;grid-template-columns:repeat(auto-fit,minmax(180px,1fr));gap:12px;margin-bottom:16px}
.kpi{background:#fff;border-radius:8px;padding:16px;box-shadow:0 1px 3px rgba(0,0,0,0.08)}
.kpi-label{font-size:11px;font-weight:600;color:#6b7280;text-transform:uppercase;letter-spacing:0.5px;margin-bottom:4px}
.kpi-value{font-size:24px;font-weight:700;color:#1f2937}
.kpi-sub{font-size:12px;margin-top:2px}
.kpi-sub.positive{color:#059669}
.kpi-sub.negative{color:#dc2626}
.chart-row{display:grid;grid-template-columns:1fr 1fr;gap:16px;margin-bottom:16px}
.chart-card{background:#fff;border-radius:8px;padding:16px;box-shadow:0 1px 3px rgba(0,0,0,0.08);min-height:340px;position:relative}
.chart-card.full{grid-column:1/-1}
.chart-title{font-size:13px;font-weight:600;color:#374151;margin-bottom:8px;text-transform:uppercase;letter-spacing:0.3px}
.chart-copy-btn{position:absolute;top:8px;right:8px;z-index:10;background:#fff;border:1px solid #d1d5db;border-radius:6px;padding:2px 7px;font-size:13px;cursor:pointer;opacity:0;transition:opacity 0.15s;line-height:1.4}
.chart-card:hover .chart-copy-btn{opacity:0.6}
.chart-copy-btn:hover{opacity:1!important;border-color:#2563eb}
.table-wrap{background:#fff;border-radius:8px;padding:16px;box-shadow:0 1px 3px rgba(0,0,0,0.08);overflow-x:auto;margin-bottom:16px}
table{width:100%;border-collapse:collapse;font-size:13px}
th{text-align:left;padding:8px 12px;border-bottom:2px solid #e5e7eb;font-weight:600;color:#6b7280;font-size:11px;text-transform:uppercase;letter-spacing:0.5px;cursor:pointer;white-space:nowrap;user-select:none}
th:hover{color:#2563eb}
td{padding:8px 12px;border-bottom:1px solid #f3f4f6;white-space:nowrap}
tr:hover td{background:#f8fafc}
tr.clickable{cursor:pointer}
tr.clickable:hover td{background:#eff6ff}
tr.total-row td{font-weight:700;background:#f8fafc;border-top:2px solid #d1d5db}
.positive{color:#059669}
.negative{color:#dc2626}
.neutral{color:#6b7280}
.badge{display:inline-block;padding:2px 8px;border-radius:10px;font-size:11px;font-weight:600}
.badge-green{background:#d1fae5;color:#065f46}
.badge-red{background:#fee2e2;color:#991b1b}
.hidden{display:none!important}
.panel{display:none}
.panel.active{display:block}
@media(max-width:900px){.chart-row{grid-template-columns:1fr}.header{flex-direction:column;text-align:center}}
.subseg-chips{display:flex;gap:6px;flex-wrap:wrap}
.subseg-chip,.view-chip{padding:4px 12px;border-radius:14px;font-size:12px;cursor:pointer;border:1px solid #d1d5db;background:#fff;color:#374151;font-weight:500;transition:all 0.15s}
.subseg-chip:hover,.view-chip:hover{border-color:#2563eb;color:#2563eb}
.subseg-chip.active{background:#2563eb;color:#fff;border-color:#2563eb}
.view-chip.active{background:#1e3a5f;color:#fff;border-color:#1e3a5f}
.align-right{text-align:right}
.upload-zone{border:2px dashed #d1d5db;border-radius:12px;padding:48px 24px;text-align:center;cursor:pointer;transition:all 0.2s;background:#fafbfc;margin-bottom:16px}
.upload-zone:hover,.upload-zone.dragover{border-color:#2563eb;background:#eff6ff}
.upload-zone h3{font-size:16px;color:#374151;margin-bottom:8px}
.upload-zone p{font-size:13px;color:#6b7280;margin:0}
.upload-zone input[type=file]{display:none}
.data-info{background:#fff;border-radius:8px;padding:20px;box-shadow:0 1px 3px rgba(0,0,0,0.08);margin-bottom:16px}
.data-info h3{font-size:14px;font-weight:600;color:#374151;margin-bottom:12px;text-transform:uppercase;letter-spacing:0.3px}
.data-info-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(180px,1fr));gap:12px}
.data-info-item{padding:12px;background:#f8fafc;border-radius:6px;border:1px solid #e5e7eb}
.data-info-item .di-label{font-size:11px;font-weight:600;color:#6b7280;text-transform:uppercase;letter-spacing:0.5px;margin-bottom:4px}
.data-info-item .di-value{font-size:16px;font-weight:700;color:#1f2937}
.status-msg{padding:12px 16px;border-radius:8px;font-size:13px;margin-bottom:12px;display:none}
.status-msg.show{display:block}
.status-msg.success{background:#d1fae5;color:#065f46}
.status-msg.error{background:#fee2e2;color:#991b1b}
.status-msg.info{background:#dbeafe;color:#1e40af}
#table-co-share-ts{font-size:11px;white-space:nowrap}
#table-co-share-ts th,#table-co-share-ts td{padding:4px 8px;min-width:60px}
#table-co-share-ts thead th:first-child{position:sticky;left:0;z-index:2;background:#f8fafc}
#table-co-share-ts tbody td:first-child{position:sticky;left:0;z-index:1;background:#fff}
.table-wrap{overflow-x:auto}
.btn{padding:8px 20px;border-radius:6px;font-size:13px;font-weight:600;cursor:pointer;border:none;transition:all 0.15s;display:inline-flex;align-items:center;gap:6px}
.btn-primary{background:#2563eb;color:#fff}.btn-primary:hover{background:#1d4ed8}
.btn-outline{background:#fff;color:#6b7280;border:1px solid #d1d5db}.btn-outline:hover{border-color:#2563eb;color:#2563eb}
.btn-danger{background:#fff;color:#dc2626;border:1px solid #fecaca}.btn-danger:hover{background:#fef2f2;border-color:#dc2626}
.spinner{display:inline-block;width:14px;height:14px;border:2px solid rgba(255,255,255,0.3);border-top-color:#fff;border-radius:50%;animation:spin 0.6s linear infinite}
@keyframes spin{to{transform:rotate(360deg)}}
.chat-setup{max-width:480px;margin:60px auto;text-align:center;background:#fff;border-radius:12px;padding:40px;box-shadow:0 2px 8px rgba(0,0,0,0.08)}
.chat-setup h3{font-size:20px;color:#1f2937;margin-bottom:8px}
.chat-setup p{font-size:13px;color:#6b7280;margin-bottom:16px}
.chat-setup input[type=password]{width:100%;padding:10px 14px;border:1px solid #d1d5db;border-radius:8px;font-size:14px;margin-bottom:12px;box-sizing:border-box}
.chat-setup .chat-help{font-size:11px;color:#9ca3af;margin-top:12px}
.chat-setup .chat-help a{color:#2563eb}
#chat-interface{display:flex;flex-direction:column;height:calc(100vh - 160px);min-height:500px}
.chat-toolbar{display:flex;gap:8px;padding:8px 0;border-bottom:1px solid #e5e7eb;margin-bottom:8px;flex-shrink:0;flex-wrap:wrap;align-items:center}
.chat-model-selector{margin-left:auto;background:#f3f4f6;border:1px solid #d1d5db;border-radius:6px;padding:3px 8px;font-size:11px;color:#374151;cursor:pointer;max-width:180px}
.chat-suggestions{display:grid;grid-template-columns:1fr 1fr;gap:10px;padding:16px 0;flex-shrink:0}
.chat-suggestion{background:#fff;border:1px solid #e5e7eb;border-radius:10px;padding:14px;cursor:pointer;text-align:left;transition:all 0.15s;font-size:12px;color:#374151;line-height:1.4}
.chat-suggestion:hover{border-color:#2563eb;background:#eff6ff}
.chat-suggestion .sug-icon{font-size:16px;margin-bottom:4px}
.chat-messages{flex:1;overflow-y:auto;padding:8px 0;display:flex;flex-direction:column;gap:12px}
.chat-msg{max-width:85%;padding:12px 16px;border-radius:12px;font-size:13px;line-height:1.6;position:relative;word-wrap:break-word}
.chat-msg.user{align-self:flex-end;background:#2563eb;color:#fff;border-bottom-right-radius:4px}
.chat-msg.assistant{align-self:flex-start;background:#fff;border:1px solid #e5e7eb;color:#1f2937;border-bottom-left-radius:4px}
.chat-msg .msg-actions{display:none;position:absolute;top:4px;right:4px;gap:4px}
.chat-msg:hover .msg-actions{display:flex}
.msg-action-btn{background:none;border:none;cursor:pointer;font-size:14px;padding:2px 4px;border-radius:4px;opacity:0.6;transition:opacity 0.15s}
.msg-action-btn:hover{opacity:1}
.msg-action-btn.saved{opacity:1;color:#f59e0b}
.chat-msg .msg-text p{margin:4px 0}
.chat-msg .msg-text b,.chat-msg .msg-text strong{font-weight:600}
.chat-msg .msg-text code{background:rgba(0,0,0,0.06);padding:1px 4px;border-radius:3px;font-size:12px;font-family:monospace}
.chat-msg .msg-text ul,.chat-msg .msg-text ol{margin:4px 0 4px 18px;padding:0}
.chat-code-result{background:#f1f5f9;border:1px solid #e2e8f0;border-radius:8px;padding:10px 14px;margin:8px 0;font-family:monospace;font-size:12px;white-space:pre-wrap;color:#334155;max-height:200px;overflow-y:auto}
.chat-code-error{background:#fef2f2;border:1px solid #fecaca;color:#991b1b}
.chat-chart-container{margin:10px 0;border-radius:8px;overflow:visible;min-height:340px;position:relative}
.chat-chart-copy{position:absolute;top:4px;right:4px;z-index:10;background:#fff;border:1px solid #d1d5db;border-radius:6px;padding:3px 8px;font-size:11px;cursor:pointer;opacity:0;transition:opacity 0.15s;color:#374151}
.chat-chart-container:hover .chat-chart-copy{opacity:0.8}
.chat-chart-copy:hover{opacity:1!important;border-color:#2563eb;color:#2563eb}
.chat-table-container{margin:8px 0;overflow-x:auto;max-height:400px;overflow-y:auto}
.chat-table-container table{font-size:11px}
.chat-input-area{display:flex;gap:8px;padding:12px 0;border-top:1px solid #e5e7eb;flex-shrink:0;align-items:flex-end}
#chat-input{flex:1;padding:10px 14px;border:1px solid #d1d5db;border-radius:10px;font-size:13px;resize:none;font-family:inherit;max-height:120px;min-height:42px;line-height:1.4}
#chat-input:focus{outline:none;border-color:#2563eb;box-shadow:0 0 0 2px rgba(37,99,235,0.15)}
.typing-indicator{align-self:flex-start;background:#fff;border:1px solid #e5e7eb;border-radius:12px;padding:12px 20px;display:flex;gap:4px}
.typing-dot{width:6px;height:6px;background:#9ca3af;border-radius:50%;animation:typingBounce 1.2s infinite}
.typing-dot:nth-child(2){animation-delay:0.2s}
.typing-dot:nth-child(3){animation-delay:0.4s}
@keyframes typingBounce{0%,60%,100%{transform:translateY(0)}30%{transform:translateY(-6px)}}
</style>
</head>
<body>
<div class="header">
  <h1>JANCHOR AUTO TRACKER</h1>
  <div class="seg-tabs" id="segTabs">
    <div class="seg-tab active" data-seg="2W">2W</div>
    <div class="seg-tab" data-seg="PV">PV</div>
    <div class="seg-tab" data-seg="3W">3W</div>
    <div class="seg-tab" data-seg="MHCV">MHCV</div>
    <div class="seg-tab" data-seg="LCV">LCV</div>
  </div>
</div>
<div class="nav" id="navTabs">
  <div class="nav-tab active" data-tab="overview">Industry Overview</div>
  <div class="nav-tab" data-tab="company">Company Deep-Dive</div>
  <div class="nav-tab" data-tab="state">State Deep-Dive</div>
  <div class="nav-tab" data-tab="zone">Zone Deep-Dive</div>
  <div class="nav-tab" data-tab="chat">&#128172; Chat</div>
  <div class="nav-tab" data-tab="data" style="margin-left:auto;color:#9ca3af">&#9881; Data</div>
</div>
<div class="main">

<!-- OVERVIEW PANEL -->
<div class="panel active" id="panel-overview">
  <div class="filter-bar">
    <label>View:</label>
    <div class="subseg-chips" id="viewChips-overview">
      <div class="view-chip active" data-view="quarterly">Quarterly</div>
      <div class="view-chip" data-view="annual">Annual</div>
    </div>
    <div class="sep"></div>
    <label>Period:</label>
    <select id="sel-period-overview"></select>
    <div class="sep"></div>
    <label>Subsegment:</label>
    <div class="subseg-chips" id="subsegChips-overview"></div>
  </div>
  <div class="kpi-row" id="kpi-overview"></div>
  <div class="chart-row">
    <div class="chart-card"><div class="chart-title" id="title-ov-vol">Industry Volume Trend</div><div id="chart-ov-vol" style="height:300px"></div></div>
    <div class="chart-card"><div class="chart-title">Market Share Trend - Top Companies</div><div id="chart-ov-share" style="height:300px"></div></div>
  </div>
  <div class="chart-row">
    <div class="chart-card full"><div class="chart-title">YoY Volume Growth Trend (%)</div><div id="chart-ov-yoy" style="height:280px"></div></div>
  </div>
  <div class="chart-row">
    <div class="chart-card"><div class="chart-title" id="title-ov-zone-split">Zone Volume Split</div><div id="chart-ov-zone-split" style="height:300px"></div></div>
    <div class="chart-card"><div class="chart-title">Zone Contribution Trend (%)</div><div id="chart-ov-zone-trend" style="height:300px"></div></div>
  </div>
  <div id="subseg-mix-overview" style="display:none">
    <div class="chart-row"><div class="chart-card" style="flex:1"><div class="chart-title" id="title-ov-subseg-mix">Subsegment Mix (%)</div><div id="chart-ov-subseg-mix" style="height:300px"></div></div></div>
    <div class="table-wrap"><div class="chart-title" id="title-ov-subseg-vol">Subsegment Volumes</div><table id="table-ov-subseg-vol"><thead></thead><tbody></tbody></table></div>
    <div class="table-wrap"><div class="chart-title" id="title-ov-subseg-yoy">Subsegment YoY Growth (%)</div><table id="table-ov-subseg-yoy"><thead></thead><tbody></tbody></table></div>
  </div>
  <div class="table-wrap">
    <div class="chart-title">Company Rankings (<span id="ov-latest-qtr"></span>)</div>
    <table id="table-ov-companies"><thead></thead><tbody></tbody></table>
  </div>
</div>

<!-- COMPANY PANEL -->
<div class="panel" id="panel-company">
  <div class="filter-bar">
    <label>Company:</label>
    <select id="sel-company"></select>
    <div class="sep"></div>
    <label>View:</label>
    <div class="subseg-chips" id="viewChips-company">
      <div class="view-chip active" data-view="quarterly">Quarterly</div>
      <div class="view-chip" data-view="annual">Annual</div>
    </div>
    <div class="sep"></div>
    <label>Period:</label>
    <select id="sel-period-company"></select>
    <div class="sep"></div>
    <label>Subsegment:</label>
    <div class="subseg-chips" id="subsegChips-company"></div>
    <div class="sep"></div>
    <label>Geo:</label>
    <div class="subseg-chips" id="geoChips-company">
      <div class="view-chip active" data-geo="state">State</div>
      <div class="view-chip" data-geo="zone">Zone</div>
    </div>
  </div>
  <div class="kpi-row" id="kpi-company"></div>
  <div class="chart-row">
    <div class="chart-card"><div class="chart-title" id="title-co-vol">Volume Trend</div><div id="chart-co-vol" style="height:300px"></div></div>
    <div class="chart-card"><div class="chart-title">Market Share Trend</div><div id="chart-co-share" style="height:300px"></div></div>
  </div>
  <div class="chart-row">
    <div class="chart-card full"><div class="chart-title">YoY Volume Growth Trend (%)</div><div id="chart-co-yoy" style="height:280px"></div></div>
  </div>
  <div id="subseg-mix-company" style="display:none">
    <div class="chart-row"><div class="chart-card" style="flex:1"><div class="chart-title" id="title-co-subseg-mix">Subsegment Mix (%)</div><div id="chart-co-subseg-mix" style="height:300px"></div></div></div>
    <div class="table-wrap"><div class="chart-title" id="title-co-subseg-vol">Subsegment Volumes</div><table id="table-co-subseg-vol"><thead></thead><tbody></tbody></table></div>
    <div class="table-wrap"><div class="chart-title" id="title-co-subseg-yoy">Subsegment YoY Growth (%)</div><table id="table-co-subseg-yoy"><thead></thead><tbody></tbody></table></div>
  </div>
  <div class="chart-row">
    <div class="chart-card"><div class="chart-title" id="title-co-states">Top States by Volume</div><div id="chart-co-states" style="height:400px"></div></div>
    <div class="chart-card"><div class="chart-title" id="title-co-share-chg">Market Share Change by State (YoY, pp)</div><div id="chart-co-share-chg" style="height:400px"></div></div>
  </div>
  <div class="chart-row">
    <div class="chart-card full"><div class="chart-title" id="title-co-contrib">Sales Contribution Trend by State (%)</div><div id="chart-co-contrib" style="height:340px"></div></div>
  </div>
  <div class="table-wrap">
    <div class="chart-title" id="title-co-share-ts">Market Share Time-Series by State (%)</div>
    <table id="table-co-share-ts"><thead></thead><tbody></tbody></table>
  </div>
  <div class="table-wrap">
    <div class="chart-title" id="title-co-geo-details">State-wise Details</div>
    <table id="table-co-states"><thead></thead><tbody></tbody></table>
  </div>
</div>

<!-- STATE PANEL -->
<div class="panel" id="panel-state">
  <div class="filter-bar">
    <label>Zone:</label>
    <select id="sel-zone"><option value="All">All Zones</option></select>
    <label style="margin-left:8px">State:</label>
    <select id="sel-state"></select>
    <div class="sep"></div>
    <label>View:</label>
    <div class="subseg-chips" id="viewChips-state">
      <div class="view-chip active" data-view="quarterly">Quarterly</div>
      <div class="view-chip" data-view="annual">Annual</div>
    </div>
    <div class="sep"></div>
    <label>Period:</label>
    <select id="sel-period-state"></select>
    <div class="sep"></div>
    <label>Subsegment:</label>
    <div class="subseg-chips" id="subsegChips-state"></div>
  </div>
  <div class="kpi-row" id="kpi-state"></div>
  <div class="chart-row">
    <div class="chart-card"><div class="chart-title" id="title-st-vol">Industry Volume Trend in State</div><div id="chart-st-vol" style="height:300px"></div></div>
    <div class="chart-card"><div class="chart-title" id="title-st-pie">Market Share Breakdown</div><div id="chart-st-pie" style="height:300px"></div></div>
  </div>
  <div class="chart-row">
    <div class="chart-card"><div class="chart-title">Market Share Trends - Top Companies</div><div id="chart-st-share" style="height:300px"></div></div>
    <div class="chart-card"><div class="chart-title">YoY Volume Growth Trend (%)</div><div id="chart-st-yoy" style="height:300px"></div></div>
  </div>
  <div id="subseg-mix-state" style="display:none">
    <div class="chart-row"><div class="chart-card" style="flex:1"><div class="chart-title" id="title-st-subseg-mix">Subsegment Mix (%)</div><div id="chart-st-subseg-mix" style="height:300px"></div></div></div>
    <div class="table-wrap"><div class="chart-title" id="title-st-subseg-vol">Subsegment Volumes</div><table id="table-st-subseg-vol"><thead></thead><tbody></tbody></table></div>
    <div class="table-wrap"><div class="chart-title" id="title-st-subseg-yoy">Subsegment YoY Growth (%)</div><table id="table-st-subseg-yoy"><thead></thead><tbody></tbody></table></div>
  </div>
  <div class="table-wrap">
    <div class="chart-title">Company Rankings in State</div>
    <table id="table-st-companies"><thead></thead><tbody></tbody></table>
  </div>
</div>

<!-- ZONE DEEP-DIVE PANEL -->
<div class="panel" id="panel-zone">
  <div class="filter-bar">
    <label>Zone:</label>
    <select id="sel-zone-tab"></select>
    <div class="sep"></div>
    <label>View:</label>
    <div class="subseg-chips" id="viewChips-zone">
      <div class="view-chip active" data-view="quarterly">Quarterly</div>
      <div class="view-chip" data-view="annual">Annual</div>
    </div>
    <div class="sep"></div>
    <label>Period:</label>
    <select id="sel-period-zone"></select>
    <div class="sep"></div>
    <label>Subsegment:</label>
    <div class="subseg-chips" id="subsegChips-zone"></div>
  </div>
  <div class="kpi-row" id="kpi-zone"></div>
  <div class="chart-row">
    <div class="chart-card"><div class="chart-title" id="title-zn-vol">Zone Volume Trend</div><div id="chart-zn-vol" style="height:300px"></div></div>
    <div class="chart-card"><div class="chart-title" id="title-zn-pie">Market Share Breakdown</div><div id="chart-zn-pie" style="height:300px"></div></div>
  </div>
  <div class="chart-row">
    <div class="chart-card"><div class="chart-title">Market Share Trends - Top Companies</div><div id="chart-zn-share" style="height:300px"></div></div>
    <div class="chart-card"><div class="chart-title">YoY Volume Growth Trend (%)</div><div id="chart-zn-yoy" style="height:300px"></div></div>
  </div>
  <div id="subseg-mix-zone" style="display:none">
    <div class="chart-row"><div class="chart-card" style="flex:1"><div class="chart-title" id="title-zn-subseg-mix">Subsegment Mix (%)</div><div id="chart-zn-subseg-mix" style="height:300px"></div></div></div>
    <div class="table-wrap"><div class="chart-title" id="title-zn-subseg-vol">Subsegment Volumes</div><table id="table-zn-subseg-vol"><thead></thead><tbody></tbody></table></div>
    <div class="table-wrap"><div class="chart-title" id="title-zn-subseg-yoy">Subsegment YoY Growth (%)</div><table id="table-zn-subseg-yoy"><thead></thead><tbody></tbody></table></div>
  </div>
  <div class="chart-row">
    <div class="chart-card"><div class="chart-title" id="title-zn-states">States by Volume</div><div id="chart-zn-states" style="height:400px"></div></div>
    <div class="chart-card"><div class="chart-title" id="title-zn-state-contrib">State Contribution Trend (%)</div><div id="chart-zn-state-contrib" style="height:400px"></div></div>
  </div>
  <div class="table-wrap">
    <div class="chart-title" id="title-zn-state-table">State-wise Details</div>
    <table id="table-zn-states"><thead></thead><tbody></tbody></table>
  </div>
  <div class="table-wrap">
    <div class="chart-title" id="title-zn-comp-table">Company Rankings in Zone</div>
    <table id="table-zn-companies"><thead></thead><tbody></tbody></table>
  </div>
</div>

<!-- DATA MANAGEMENT PANEL -->
<!-- CHAT PANEL -->
<div class="panel" id="panel-chat">
  <div id="chat-setup" class="chat-setup">
    <h3>&#129302; Chat with Your Data</h3>
    <p>Ask questions, create charts, run calculations on your auto industry data using AI.</p>
    <label style="font-weight:600;margin-bottom:4px;display:block;font-size:13px">Select Model:</label>
    <select id="sel-chat-model" style="width:100%;padding:10px 14px;border:1px solid #d1d5db;border-radius:8px;font-size:14px;margin-bottom:12px;background:#fff;cursor:pointer"></select>
    <input type="password" id="api-key-input" placeholder="Enter your API key">
    <button class="btn btn-primary" id="btn-save-key" style="width:100%">Save &amp; Start Chatting</button>
    <p class="chat-help" id="chat-key-help">Get your API key from <a id="chat-key-link" href="https://console.anthropic.com/settings/keys" target="_blank">console.anthropic.com</a></p>
  </div>
  <div id="chat-interface" style="display:none">
    <div class="chat-toolbar">
      <button class="btn btn-outline" id="btn-chat-clear" style="font-size:12px;padding:4px 12px">&#128465; Clear</button>
      <button class="btn btn-outline" id="btn-chat-export" style="font-size:12px;padding:4px 12px">&#128190; Export Saved</button>
      <button class="btn btn-outline" id="btn-chat-remove-key" style="font-size:12px;padding:4px 12px">&#128274; Remove Key</button>
      <button class="btn btn-outline" id="btn-chat-download-data" style="font-size:12px;padding:4px 12px" title="Download all data as CSV for use with any AI tool">&#128229; Download Data CSV</button>
      <select id="sel-toolbar-model" class="chat-model-selector"></select>
    </div>
    <div class="chat-suggestions" id="chat-suggestions">
      <div class="chat-suggestion" data-prompt="What are the top 5 companies by market share in the latest quarter? Show a table.">
        <div class="sug-icon">&#128202;</div>Top 5 companies by market share in the latest quarter
      </div>
      <div class="chat-suggestion" data-prompt="Plot the market share trend for the top 3 companies over the last 3 years as a line chart.">
        <div class="sug-icon">&#128200;</div>Market share trend for top 3 companies (line chart)
      </div>
      <div class="chat-suggestion" data-prompt="Which states have seen the biggest market share gain and loss in the latest year vs previous year? Show top 5 gainers and losers.">
        <div class="sug-icon">&#127919;</div>States with biggest share gain/loss vs last year
      </div>
      <div class="chat-suggestion" data-prompt="Calculate the industry CAGR from FY18 to FY25 and tell me which company has grown fastest.">
        <div class="sug-icon">&#128640;</div>Industry CAGR &amp; fastest growing company
      </div>
    </div>
    <div class="chat-messages" id="chat-messages"></div>
    <div class="chat-input-area">
      <textarea id="chat-input" placeholder="Ask about your data... (Enter to send, Shift+Enter for new line)" rows="1"></textarea>
      <button class="btn btn-primary" id="btn-chat-send">Send</button>
    </div>
  </div>
</div>

<div class="panel" id="panel-data">
  <div style="max-width:800px;margin:0 auto">
    <div class="status-msg" id="data-status"></div>
    <div class="data-info" id="data-info-box">
      <h3>Current Data</h3>
      <div class="data-info-grid" id="data-info-grid"></div>
    </div>
    <div class="upload-zone" id="upload-zone">
      <input type="file" id="file-input" accept=".xlsx,.xls">
      <h3>&#128193; Upload New Data File</h3>
      <p>Drag &amp; drop your Excel file here, or click to browse</p>
      <p style="margin-top:8px;font-size:11px;color:#9ca3af">Accepts .xlsx files with raw data sheets (PVs, 2Ws, 3Ws, M&amp;HCVs, LCVs)</p>
    </div>
    <div style="display:flex;gap:12px;margin-top:12px;flex-wrap:wrap">
      <button class="btn btn-danger" id="btn-reset-data" style="display:none">Reset to Embedded Data</button>
    </div>
  </div>
</div>

</div>

<script>
// ============================================
// DATA
// ============================================
const EMBEDDED_DATA = ''' + data_json + ''';
const RAW = (function(){try{const s=localStorage.getItem("janchor_auto_data");if(s)return JSON.parse(s);}catch(e){}return EMBEDDED_DATA;})();
const DATA_IS_CUSTOM = (RAW !== EMBEDDED_DATA);
const Q = RAW.quarters;
const NQ = Q.length;
const ROWS = RAW.rows;

// ============================================
// FISCAL YEAR / ANNUAL HELPERS
// ============================================
const FYS = []; // ['FY17','FY18',...] derived dynamically from Q
const FY_Q_IDXS = {}; // {'FY17':[0,1,2,3], ...}
(function() {
  for (let i = 0; i < Q.length; i++) {
    const fy = 'FY' + Q[i].slice(4); // "Q1FY17" -> "FY17"
    if (!FY_Q_IDXS[fy]) { FYS.push(fy); FY_Q_IDXS[fy] = []; }
    FY_Q_IDXS[fy].push(i);
  }
})();
const NFY = FYS.length;

function annualVols(qVols) {
  return FYS.map(fy => FY_Q_IDXS[fy].reduce((s, qi) => s + qVols[qi], 0));
}

function fyDates() {
  return FYS.map(fy => {
    const yr = 2000 + parseInt(fy.slice(2)) - 1;
    return new Date(yr, 3, 1); // April of start year
  });
}
const FYDATES = fyDates();
const FYLABELS = [...FYS]; // "FY17", "FY18", ...

// ============================================
// CONSTANTS & CONFIG
// ============================================
const COMPANY_COLORS = {
  'Maruti':'#FF6B35','Hyundai':'#004B87','Tata Motors':'#1B1464','M&M':'#C62828','Kia':'#BB162B',
  'Toyota':'#EB0A1E','Honda':'#CC0000','Renault':'#F4B400','Skoda':'#4CAF50','MG Motor':'#FF5722',
  'Volkswagen':'#1565C0','Nissan':'#C0392B','Ford':'#1A237E','Force Motors':'#795548',
  'Hero':'#00A651','Bajaj':'#0055A5','Royal Enfield':'#212121','TVS Motors':'#E30613',
  'Yamaha':'#005AAC','Suzuki':'#003DA5','Ather':'#00BCD4','Okinawa':'#FF9800',
  'Piaggio':'#4CAF50','Ashok Leyland':'#C41230','VECV':'#FFB800','SML Isuzu':'#607D8B',
  'Atul Auto':'#8BC34A','TI Clean Mobility':'#00ACC1','Continental Engines':'#9C27B0',
  'Scooters India':'#CDDC39','Isuzu':'#455A64',
  'AMW Motors':'#3F51B5','Pinnacle Mobility':'#009688','PMI Electro Mobility':'#E91E63',
  'Volvo Group':'#1B5E20','VECV-Volvo':'#33691E','Maruti Suzuki':'#FF6B35','Isuzu Motors':'#455A64'
};
const PALETTE = ['#2563eb','#dc2626','#059669','#d97706','#7c3aed','#db2777','#0891b2','#65a30d','#ea580c','#4f46e5','#be123c','#0d9488','#ca8a04','#9333ea','#e11d48'];
function getColor(company, idx) { return COMPANY_COLORS[company] || PALETTE[idx % PALETTE.length]; }

// ============================================
// STATE
// ============================================
let currentSegment = '2W';
let currentTab = 'overview';
let currentSubseg = 'All';
let currentCompany = '';
let currentZone = 'All';
let currentState = '';
let currentGeo = 'state'; // 'state' or 'zone' — for company deep-dive geo breakdown
let currentZoneTab = ''; // selected zone on the Zone Deep-Dive tab
let viewMode = 'quarterly'; // per-tab
const viewModes = {overview:'quarterly', company:'quarterly', state:'quarterly', zone:'quarterly', chat:'quarterly'};
const selectedPeriods = {overview: NQ-1, company: NQ-1, state: NQ-1, zone: NQ-1, chat: NQ-1};

// Chat state
let chatHistory = []; // [{id, role, content, timestamp, saved, renderedHTML}]
var CHAT_MODELS = {
  'claude-sonnet': { provider:'anthropic', name:'Claude Sonnet', model:'claude-sonnet-4-20250514', keyPrefix:'sk-ant-', keyPlaceholder:'Enter Anthropic API key (sk-ant-...)', keyLink:'https://console.anthropic.com/settings/keys', keyLinkLabel:'console.anthropic.com' },
  'gemini-flash': { provider:'google', name:'Gemini 2.5 Flash', model:'gemini-2.5-flash', keyPrefix:'AIza', keyPlaceholder:'Enter Google AI API key (AIza...)', keyLink:'https://aistudio.google.com/apikey', keyLinkLabel:'aistudio.google.com' },
  'gemini-pro': { provider:'google', name:'Gemini 2.5 Pro', model:'gemini-2.5-pro', keyPrefix:'AIza', keyPlaceholder:'Enter Google AI API key (AIza...)', keyLink:'https://aistudio.google.com/apikey', keyLinkLabel:'aistudio.google.com' },
  'gemini3-flash': { provider:'google', name:'Gemini 3 Flash', model:'gemini-3-flash-preview', keyPrefix:'AIza', keyPlaceholder:'Enter Google AI API key (AIza...)', keyLink:'https://aistudio.google.com/apikey', keyLinkLabel:'aistudio.google.com' },
  'gemini3-pro': { provider:'google', name:'Gemini 3.1 Pro', model:'gemini-3.1-pro-preview', keyPrefix:'AIza', keyPlaceholder:'Enter Google AI API key (AIza...)', keyLink:'https://aistudio.google.com/apikey', keyLinkLabel:'aistudio.google.com' }
};
var selectedModel = 'claude-sonnet';
try { selectedModel = localStorage.getItem('janchor_selected_model') || 'claude-sonnet'; } catch(e) {}
if (!CHAT_MODELS[selectedModel]) selectedModel = 'claude-sonnet';

// Per-provider API keys
var chatApiKeys = {};
try { chatApiKeys = JSON.parse(localStorage.getItem('janchor_api_keys') || '{}'); } catch(e) { chatApiKeys = {}; }
// Migrate old single key
(function() {
  var oldKey = null;
  try { oldKey = localStorage.getItem('janchor_api_key'); } catch(e) {}
  if (oldKey && !chatApiKeys['anthropic']) {
    chatApiKeys['anthropic'] = oldKey;
    try { localStorage.setItem('janchor_api_keys', JSON.stringify(chatApiKeys)); localStorage.removeItem('janchor_api_key'); } catch(e) {}
  }
})();
function getChatApiKey() { return chatApiKeys[CHAT_MODELS[selectedModel].provider] || ''; }
try { chatHistory = JSON.parse(localStorage.getItem('janchor_chat_history') || '[]'); } catch(e) { chatHistory = []; }

function getViewMode() { return viewModes[currentTab]; }
function getSelectedPeriod() { return selectedPeriods[currentTab]; }

// ============================================
// INDEXES
// ============================================
let segRows = [], segCompanies = [], segStates = [], segZones = [], segSubsegs = [];
let zoneStateMap = {};

function buildIndexes() {
  segRows = [];
  const compSet = new Set(), stateSet = new Set(), zoneSet = new Set(), subsegSet = new Set();
  zoneStateMap = {};
  for (let i = 0; i < ROWS.length; i++) {
    if (ROWS[i][0] === currentSegment) {
      segRows.push(i);
      compSet.add(ROWS[i][4]); stateSet.add(ROWS[i][3]); zoneSet.add(ROWS[i][2]); subsegSet.add(ROWS[i][1]);
      const z = ROWS[i][2], s = ROWS[i][3];
      if (!zoneStateMap[z]) zoneStateMap[z] = new Set();
      zoneStateMap[z].add(s);
    }
  }
  segCompanies = [...compSet].sort();
  segStates = [...stateSet].sort();
  segZones = [...zoneSet].sort();
  segSubsegs = [...subsegSet].sort();
  for (const z in zoneStateMap) zoneStateMap[z] = [...zoneStateMap[z]].sort();
}

// ============================================
// DATA HELPERS
// ============================================
function filterRows(company, state, subseg, zone) {
  let rows = segRows;
  if (subseg && subseg !== 'All') rows = rows.filter(i => ROWS[i][1] === subseg);
  if (company) rows = rows.filter(i => ROWS[i][4] === company);
  if (state) rows = rows.filter(i => ROWS[i][3] === state);
  if (zone) rows = rows.filter(i => ROWS[i][2] === zone);
  return rows;
}

function sumVolumes(rowIdxs) {
  const vols = new Array(NQ).fill(0);
  for (const i of rowIdxs) { for (let q = 0; q < NQ; q++) vols[q] += (ROWS[i][5+q]||0); }
  return vols;
}

function getIndustryVols(sub) { return sumVolumes(filterRows(null,null,sub)); }
function getCompanyVols(co,sub) { return sumVolumes(filterRows(co,null,sub)); }
function getStateIndustryVols(st,sub) { return sumVolumes(filterRows(null,st,sub)); }
function getStateCompanyVols(st,co,sub) { return sumVolumes(filterRows(co,st,sub)); }
function getZoneIndustryVols(zone,sub) { return sumVolumes(filterRows(null,null,sub,zone)); }
function getZoneCompanyVols(zone,co,sub) { return sumVolumes(filterRows(co,null,sub,zone)); }
function computeShare(cv,iv) { return cv.map((v,i) => iv[i]>0 ? v/iv[i]*100 : 0); }

// Get volumes for current view mode
function tsVols(qVols) { return getViewMode()==='annual' ? annualVols(qVols) : qVols; }
function tsDates() {
  if (getViewMode()==='annual') return FYS.map(fy => fy + (FY_Q_IDXS[fy].length < 4 ? '*' : ''));
  return QLABELS;
}
function tsLabels() {
  if (getViewMode()==='annual') return FYS.map(fy => fy + (FY_Q_IDXS[fy].length < 4 ? '*' : ''));
  return QLABELS;
}
function tsLen() { return getViewMode()==='annual' ? NFY : NQ; }

// Period helpers: convert selected period index to quarter indices for data lookup
function periodQIdxs() {
  const vm = getViewMode(), pi = getSelectedPeriod();
  if (vm === 'annual') return FY_Q_IDXS[FYS[pi]] || [];
  return [pi];
}
function periodVol(qVols) { return periodQIdxs().reduce((s,qi) => s + qVols[qi], 0); }
function periodLabel() {
  const vm = getViewMode(), pi = getSelectedPeriod();
  if (vm === 'annual') return FYS[pi] + (FY_Q_IDXS[FYS[pi]].length < 4 ? ' YTD' : '');
  return QLABELS[pi];
}
// YoY comparison period
function yoyPeriodQIdxs() {
  const vm = getViewMode(), pi = getSelectedPeriod();
  if (vm === 'annual') {
    const prevFY = pi > 0 ? FYS[pi-1] : null;
    if (!prevFY) return [];
    // For partial year, compare same quarters
    const curQs = FY_Q_IDXS[FYS[pi]];
    const prevQs = FY_Q_IDXS[prevFY];
    return prevQs.slice(0, curQs.length);
  }
  return pi >= 4 ? [pi - 4] : [];
}
function yoyPeriodVol(qVols) { return yoyPeriodQIdxs().reduce((s,qi) => s + qVols[qi], 0); }
function yoyPeriodLabel() {
  const vm = getViewMode(), pi = getSelectedPeriod();
  if (vm === 'annual') {
    if (pi > 0) {
      const curQs = FY_Q_IDXS[FYS[pi]];
      return FYS[pi-1] + (curQs.length < 4 ? ' (comparable)' : '');
    }
    return '-';
  }
  return pi >= 4 ? QLABELS[pi-4] : '-';
}

function fmt(n) {
  if (n === 0) return '0';
  if (Math.abs(n) >= 1e6) return (n/1e6).toFixed(1)+'M';
  if (Math.abs(n) >= 1e4) return (n/1e3).toFixed(0)+'K';
  if (Math.abs(n) >= 1e3) return (n/1e3).toFixed(1)+'K';
  return Math.round(n).toLocaleString('en-IN');
}
function fmtPct(n) { return n.toFixed(1)+'%'; }
function fmtPP(n) { return (n>=0?'+':'')+n.toFixed(1)+' pp'; }

function quarterDates() {
  return Q.map(q => {
    const qn=parseInt(q[1]), fy=parseInt(q.slice(4));
    const months={1:3,2:6,3:9,4:0};
    const yr=2000+fy-1+(qn===4?1:0);
    return new Date(yr,months[qn],1);
  });
}
const QDATES = quarterDates();
const QLABELS = Q.map(q => q[1]+'Q'+q.slice(2)); // "1QFY17","2QFY17",...

// Top N companies by volume in selected period
function topCompanies(n, sub, state, zone) {
  const volMap = {};
  const rows = filterRows(null, state, sub, zone);
  for (const i of rows) {
    const c = ROWS[i][4];
    volMap[c] = (volMap[c]||0) + periodQIdxs().reduce((s,qi) => s+(ROWS[i][5+qi]||0), 0);
  }
  return Object.entries(volMap).sort((a,b)=>b[1]-a[1]).slice(0,n).map(e=>e[0]);
}

// YoY growth time-series
function yoyGrowthSeries(qVols) {
  const vm = getViewMode();
  if (vm === 'annual') {
    return FYS.map((fy, i) => {
      if (i === 0) return null;
      const curQs = FY_Q_IDXS[fy];
      const prevQs = FY_Q_IDXS[FYS[i-1]];
      // FYTD: compare same number of quarters for incomplete years
      const nQ = curQs.length;
      const curVol = curQs.reduce((s, qi) => s + qVols[qi], 0);
      const prevVol = prevQs.slice(0, nQ).reduce((s, qi) => s + qVols[qi], 0);
      return prevVol > 0 ? ((curVol/prevVol-1)*100) : null;
    });
  }
  return qVols.map((v,i) => {
    if (i < 4) return null;
    const prev = qVols[i-4];
    return prev > 0 ? ((v/prev-1)*100) : null;
  });
}

// ============================================
// CHART RENDERING
// ============================================
const PLOTLY_LAYOUT = {
  margin:{l:80,r:25,t:10,b:75},
  font:{family:'-apple-system,BlinkMacSystemFont,Segoe UI,Roboto,sans-serif',size:11},
  hovermode:'x unified',
  legend:{orientation:'h',y:-0.22,x:0.5,xanchor:'center',font:{size:10}},
  xaxis:{showgrid:false,tickangle:90},
  yaxis:{showgrid:true,gridcolor:'#f0f0f0',automargin:true,title:{standoff:10}},
  plot_bgcolor:'white',paper_bgcolor:'white'
};
const PLOTLY_CONFIG = {responsive:true,displayModeBar:false};

function addChartCopyBtn(chartId) {
  const plotDiv = document.getElementById(chartId);
  if (!plotDiv) return;
  const card = plotDiv.closest('.chart-card') || plotDiv.parentElement;
  if (card.querySelector('.chart-copy-btn')) return; // already has one
  card.style.position = 'relative';
  const btn = document.createElement('button');
  btn.className = 'chart-copy-btn';
  btn.innerHTML = '&#128247;';
  btn.title = 'Copy chart as image';
  btn.onclick = function(e) {
    e.stopPropagation();
    Plotly.toImage(chartId, {format:'png', width:1000, height:550}).then(function(url){
      fetch(url).then(function(r){return r.blob();}).then(function(blob){
        try {
          navigator.clipboard.write([new ClipboardItem({'image/png':blob})]);
          btn.textContent='\\u2713'; setTimeout(function(){btn.innerHTML='&#128247;';},1500);
        } catch(err){ var w=window.open(''); w.document.write('<img src="'+url+'">'); }
      });
    });
  };
  card.appendChild(btn);
}

function plotLines(id, traces, yTitle, pctFmt) {
  const layout = {...PLOTLY_LAYOUT,
    yaxis:{...PLOTLY_LAYOUT.yaxis, title:{text:yTitle,standoff:12}, tickformat:pctFmt?'.1f':',', ticksuffix:pctFmt?'%':'', automargin:true}
  };
  Plotly.newPlot(id, traces, layout, PLOTLY_CONFIG);
  addChartCopyBtn(id);
}

function plotHBar(id, labels, values, colors) {
  const trace = {
    type:'bar', orientation:'h',
    y:labels, x:values,
    marker:{color:colors||'#2563eb'},
    text:values.map(v => typeof v==='number'?(Math.abs(v)<10?v.toFixed(1):fmt(v)):v),
    textposition:'outside', textfont:{size:10},
    hovertemplate:'%{y}: %{x:,.0f}<extra></extra>'
  };
  const layout = {
    ...PLOTLY_LAYOUT, hovermode:'closest',
    margin:{l:130,r:70,t:10,b:30},
    yaxis:{autorange:'reversed',showgrid:false,type:'category',automargin:true},
    xaxis:{showgrid:true,gridcolor:'#f0f0f0',type:'linear'}
  };
  Plotly.newPlot(id,[trace],layout,PLOTLY_CONFIG);
  addChartCopyBtn(id);
}

function plotHBarDiverging(id, labels, values, colors) {
  const trace = {
    type:'bar', orientation:'h',
    y:labels, x:values,
    marker:{color:colors},
    text:values.map(v => (v>=0?'+':'')+v.toFixed(1)),
    textposition:'outside', textfont:{size:10},
    hovertemplate:'%{y}: %{x:.1f} pp<extra></extra>'
  };
  const layout = {
    ...PLOTLY_LAYOUT, hovermode:'closest',
    margin:{l:130,r:60,t:10,b:30},
    yaxis:{autorange:'reversed',showgrid:false,type:'category',automargin:true},
    xaxis:{showgrid:true,gridcolor:'#f0f0f0',type:'linear',zeroline:true,zerolinecolor:'#94a3b8'}
  };
  Plotly.newPlot(id,[trace],layout,PLOTLY_CONFIG);
  addChartCopyBtn(id);
}

function plotDonut(id, labels, values, colors) {
  const trace = {
    type:'pie',labels,values,hole:0.45,marker:{colors},
    textinfo:'label+percent',textposition:'outside',textfont:{size:10},
    hovertemplate:'%{label}: %{value:,.0f} (%{percent})<extra></extra>',sort:false
  };
  Plotly.newPlot(id,[trace],{...PLOTLY_LAYOUT,showlegend:false,margin:{l:10,r:10,t:10,b:10}},PLOTLY_CONFIG);
  addChartCopyBtn(id);
}

function plotYoYGrowth(id, traces, title) {
  const layout = {
    ...PLOTLY_LAYOUT,
    yaxis:{...PLOTLY_LAYOUT.yaxis, title:{text:'YoY Growth (%)',standoff:12}, ticksuffix:'%', zeroline:true, zerolinecolor:'#94a3b8', zerolinewidth:1.5, automargin:true},
    shapes:[{type:'line',x0:0,x1:1,xref:'paper',y0:0,y1:0,line:{color:'#94a3b8',width:1.5,dash:'dot'}}]
  };
  Plotly.newPlot(id, traces, layout, PLOTLY_CONFIG);
  addChartCopyBtn(id);
}

// ============================================
// PERIOD SELECTOR
// ============================================
function populatePeriodSelector(tabName) {
  const sel = document.getElementById('sel-period-' + tabName);
  if (!sel) return;
  const vm = viewModes[tabName];
  sel.innerHTML = '';
  if (vm === 'annual') {
    FYS.forEach((fy,i) => {
      const lbl = fy + (FY_Q_IDXS[fy].length < 4 ? ' (YTD)' : '');
      sel.innerHTML += '<option value="'+i+'"'+(i===selectedPeriods[tabName]?' selected':'')+'>'+lbl+'</option>';
    });
  } else {
    Q.forEach((q,i) => {
      sel.innerHTML += '<option value="'+i+'"'+(i===selectedPeriods[tabName]?' selected':'')+'>'+QLABELS[i]+'</option>';
    });
  }
  sel.value = selectedPeriods[tabName];
}

// ============================================
// OVERVIEW
// ============================================
function renderOverview() {
  const sub = currentSubseg;
  const indQ = getIndustryVols(sub);

  // KPIs use selected period
  const curVol = periodVol(indQ);
  const yoyV = yoyPeriodVol(indQ);
  const yoyG = yoyV > 0 ? ((curVol/yoyV-1)*100) : 0;

  const comps = topCompanies(50,sub,null);
  let bestGainer='',bestGain=-999,bestLoser='',bestLoss=999;
  for (const c of comps) {
    const cv = getCompanyVols(c,sub);
    const curC = periodVol(cv), yoyC = yoyPeriodVol(cv);
    const sCur = curVol>0?curC/curVol*100:0;
    const sYoy = yoyV>0?yoyC/yoyV*100:0;
    const chg = sCur-sYoy;
    if (chg>bestGain && sCur>0.5){bestGain=chg;bestGainer=c;}
    if (chg<bestLoss && sYoy>0.5){bestLoss=chg;bestLoser=c;}
  }

  document.getElementById('kpi-overview').innerHTML = `
    <div class="kpi"><div class="kpi-label">Industry Volume (${periodLabel()})</div><div class="kpi-value">${fmt(curVol)}</div></div>
    <div class="kpi"><div class="kpi-label">YoY Growth vs ${yoyPeriodLabel()}</div><div class="kpi-value ${yoyG>=0?'positive':'negative'}">${yoyG>=0?'+':''}${yoyG.toFixed(1)}%</div></div>
    <div class="kpi"><div class="kpi-label">Top Share Gainer</div><div class="kpi-value" style="font-size:18px">${bestGainer}</div><div class="kpi-sub positive">${fmtPP(bestGain)}</div></div>
    <div class="kpi"><div class="kpi-label">Top Share Loser</div><div class="kpi-value" style="font-size:18px">${bestLoser}</div><div class="kpi-sub negative">${fmtPP(bestLoss)}</div></div>
  `;

  document.getElementById('ov-latest-qtr').textContent = periodLabel();
  document.getElementById('title-ov-vol').textContent = 'Industry Volume Trend (' + (getViewMode()==='annual'?'Annual':'Quarterly') + ')';

  // Volume trend
  const tsV = tsVols(indQ);
  plotLines('chart-ov-vol',[{
    x:tsDates(),y:tsV,type:'scatter',mode:'lines+markers',
    name:'Industry Volume',line:{color:'#2563eb',width:2.5},marker:{size:4},
    fill:'tozeroy',fillcolor:'rgba(37,99,235,0.08)',
    hovertemplate:'%{x}: %{y:,.0f}<extra></extra>'
  }],'Volume (units)',false);

  // Market share trend
  const top8 = topCompanies(8,sub,null);
  const shareTraces = top8.map((c,ci) => {
    const cv = getCompanyVols(c,sub);
    const share = computeShare(tsVols(cv), tsV);
    return {x:tsDates(),y:share,type:'scatter',mode:'lines',name:c,line:{color:getColor(c,ci),width:2},
      hovertemplate:c+': %{y:.1f}%<extra></extra>'};
  });
  plotLines('chart-ov-share',shareTraces,'Market Share (%)',true);

  // YoY Growth Trend
  const indYoYOv = yoyGrowthSeries(indQ);
  const top3Ov = topCompanies(3, sub, null);
  const yoyTracesOv = [
    {x:tsDates(),y:indYoYOv,type:'scatter',mode:'lines+markers',name:'Industry',line:{color:'#2563eb',width:2.5},marker:{size:4},connectgaps:false,hovertemplate:'Industry: %{y:.1f}%<extra></extra>'}
  ];
  top3Ov.forEach((c,ci) => {
    const yoy = yoyGrowthSeries(getCompanyVols(c, sub));
    yoyTracesOv.push({x:tsDates(),y:yoy,type:'scatter',mode:'lines',name:c,line:{color:getColor(c,ci+1),width:1.5},connectgaps:false,hovertemplate:c+': %{y:.1f}%<extra></extra>'});
  });
  plotYoYGrowth('chart-ov-yoy', yoyTracesOv);

  // Zone Volume Split (donut)
  document.getElementById('title-ov-zone-split').textContent = 'Zone Volume Split (' + periodLabel() + ')';
  const pQsOv = periodQIdxs();
  const zoneLabels=[], zoneVals=[], zoneColors=[];
  const ZONE_COLORS = ['#2563eb','#dc2626','#059669','#d97706','#7c3aed','#db2777','#0891b2','#65a30d','#ea580c','#4f46e5'];
  segZones.forEach((z,zi) => {
    const zv = getZoneIndustryVols(z, sub);
    const vol = pQsOv.reduce((s,qi)=>s+(zv[qi]||0),0);
    if (vol > 0) {
      zoneLabels.push(z); zoneVals.push(vol); zoneColors.push(ZONE_COLORS[zi % ZONE_COLORS.length]);
    }
  });
  plotDonut('chart-ov-zone-split', zoneLabels, zoneVals, zoneColors);

  // Zone Contribution Trend (lines)
  const znContribTraces = [];
  segZones.forEach((z,zi) => {
    const zv = getZoneIndustryVols(z, sub);
    const zTS = tsVols(zv);
    const contribTS = zTS.map((v,t) => tsV[t]>0 ? v/tsV[t]*100 : 0);
    znContribTraces.push({
      x:tsDates(), y:contribTS, type:'scatter', mode:'lines+markers',
      name:z, line:{color:ZONE_COLORS[zi % ZONE_COLORS.length],width:2}, marker:{size:3},
      hovertemplate:z+': %{y:.1f}%<extra></extra>'
    });
  });
  plotLines('chart-ov-zone-trend', znContribTraces, 'Contribution (%)', true);

  // Subsegment mix section (PV: Cars/UVs, 2W: Motorcycle/Scooters)
  var actualSubsegsOv = segSubsegs.filter(function(s){return s!=='All';});
  var mixElOv = document.getElementById('subseg-mix-overview');
  if (actualSubsegsOv.length >= 2 && sub === 'All') {
    mixElOv.style.display = 'block';
    var segNameOv = currentSegment;
    var subColorsOv = ['#2563eb','#f59e0b','#10b981','#ef4444','#8b5cf6'];
    var subVolsOv = actualSubsegsOv.map(function(s){return tsVols(getIndustryVols(s));});
    var mixTotalsOv = subVolsOv[0].map(function(_,t){return subVolsOv.reduce(function(sum,sv){return sum+sv[t];},0);});
    var subPctsOv = subVolsOv.map(function(sv){return sv.map(function(v,t){return mixTotalsOv[t]>0?v/mixTotalsOv[t]*100:0;});});
    var datesOv = tsDates(), tsLenOv = datesOv.length;

    document.getElementById('title-ov-subseg-mix').textContent = segNameOv + ' Subsegment Mix (%) - National';
    var mixTracesOv = actualSubsegsOv.map(function(s,i){return {
      x:datesOv, y:subPctsOv[i], customdata:subVolsOv[i], type:'bar', name:s,
      marker:{color:subColorsOv[i%subColorsOv.length]},
      hovertemplate:s+': %{y:.1f}%<br>Volume: %{customdata:,.0f}<extra></extra>'
    };});
    Plotly.newPlot('chart-ov-subseg-mix',mixTracesOv,{
      ...PLOTLY_LAYOUT, barmode:'stack',
      margin:{l:60,r:25,t:10,b:75},
      yaxis:{...PLOTLY_LAYOUT.yaxis, title:{text:'% Mix',standoff:10}, range:[0,100.5]},
      legend:{orientation:'h',y:-0.22,x:0.5,xanchor:'center',font:{size:10}}
    },PLOTLY_CONFIG);
    addChartCopyBtn('chart-ov-subseg-mix');

    document.getElementById('title-ov-subseg-vol').textContent = segNameOv + ' Subsegment Volumes - National';
    var volHeadOv = '<tr><th>Subsegment</th>' + datesOv.map(function(d){return '<th class="align-right">'+d+'</th>';}).join('') + '</tr>';
    var volBodyOv = '';
    actualSubsegsOv.forEach(function(s,i){
      volBodyOv += '<tr><td><b>'+s+'</b></td>';
      for(var t=0;t<tsLenOv;t++) volBodyOv += '<td class="align-right">'+fmt(subVolsOv[i][t])+'</td>';
      volBodyOv += '</tr>';
    });
    volBodyOv += '<tr class="total-row"><td><b>Total '+segNameOv+'</b></td>';
    for(var t=0;t<tsLenOv;t++) volBodyOv += '<td class="align-right"><b>'+fmt(mixTotalsOv[t])+'</b></td>';
    volBodyOv += '</tr>';
    actualSubsegsOv.forEach(function(s,i){
      volBodyOv += '<tr style="color:#6b7280;font-size:12px"><td><i>'+s+' %</i></td>';
      for(var t=0;t<tsLenOv;t++) volBodyOv += '<td class="align-right"><i>'+subPctsOv[i][t].toFixed(1)+'%</i></td>';
      volBodyOv += '</tr>';
    });
    document.querySelector('#table-ov-subseg-vol thead').innerHTML = volHeadOv;
    document.querySelector('#table-ov-subseg-vol tbody').innerHTML = volBodyOv;

    document.getElementById('title-ov-subseg-yoy').textContent = segNameOv + ' Subsegment YoY Growth (%) - National';
    var yoyHeadOv = '<tr><th>Subsegment</th>' + datesOv.map(function(d){return '<th class="align-right">'+d+'</th>';}).join('') + '</tr>';
    var yoyBodyOv = '';
    actualSubsegsOv.forEach(function(s){
      var yoyS = yoyGrowthSeries(getIndustryVols(s));
      yoyBodyOv += '<tr><td><b>'+s+'</b></td>';
      for(var t=0;t<tsLenOv;t++){
        var g=yoyS[t];
        if(g===null||g===undefined) yoyBodyOv+='<td class="align-right">-</td>';
        else yoyBodyOv+='<td class="align-right '+(g>=0?'positive':'negative')+'">'+(g>=0?'+':'')+g.toFixed(1)+'%</td>';
      }
      yoyBodyOv += '</tr>';
    });
    var overallYoYOv = yoyGrowthSeries(indQ);
    yoyBodyOv += '<tr class="total-row"><td><b>Overall '+segNameOv+'</b></td>';
    for(var t=0;t<tsLenOv;t++){
      var g2=overallYoYOv[t];
      if(g2===null||g2===undefined) yoyBodyOv+='<td class="align-right">-</td>';
      else yoyBodyOv+='<td class="align-right '+(g2>=0?'positive':'negative')+'"><b>'+(g2>=0?'+':'')+g2.toFixed(1)+'%</b></td>';
    }
    yoyBodyOv += '</tr>';
    document.querySelector('#table-ov-subseg-yoy thead').innerHTML = yoyHeadOv;
    document.querySelector('#table-ov-subseg-yoy tbody').innerHTML = yoyBodyOv;
  } else {
    mixElOv.style.display = 'none';
  }

  // Company table
  renderOverviewTable(sub);
}

function renderOverviewTable(sub) {
  const indQ = getIndustryVols(sub);
  const curIndVol = periodVol(indQ);
  const yoyIndVol = yoyPeriodVol(indQ);
  const comps = topCompanies(50,sub,null);

  const tableData = comps.map(c => {
    const cv = getCompanyVols(c,sub);
    const vol = periodVol(cv);
    const yv = yoyPeriodVol(cv);
    const growth = yv>0?((vol/yv-1)*100):0;
    const share = curIndVol>0?(vol/curIndVol*100):0;
    const shareY = yoyIndVol>0?(yv/yoyIndVol*100):0;
    return {company:c,vol,yoyVol:yv,growth,share,shareChg:share-shareY};
  }).filter(d=>d.vol>0);
  tableData.sort((a,b)=>b.vol-a.vol);

  // Total
  const totVol = tableData.reduce((s,d)=>s+d.vol,0);
  const totYoy = tableData.reduce((s,d)=>s+d.yoyVol,0);
  const totGrowth = totYoy>0?((totVol/totYoy-1)*100):0;

  const thead = document.querySelector('#table-ov-companies thead');
  thead.innerHTML = '<tr><th>#</th><th>Company</th><th class="align-right">Volume</th><th class="align-right">YoY Volume</th><th class="align-right">YoY Growth</th><th class="align-right">Mkt Share</th><th class="align-right">Share Chg</th></tr>';

  const tbody = document.querySelector('#table-ov-companies tbody');
  tbody.innerHTML = tableData.map((d,i) => `
    <tr class="clickable" onclick="drillToCompany('${esc(d.company)}')">
      <td>${i+1}</td><td><b>${d.company}</b></td>
      <td class="align-right">${fmt(d.vol)}</td><td class="align-right">${fmt(d.yoyVol)}</td>
      <td class="align-right ${d.growth>=0?'positive':'negative'}">${d.growth>=0?'+':''}${d.growth.toFixed(1)}%</td>
      <td class="align-right">${d.share.toFixed(1)}%</td>
      <td class="align-right"><span class="badge ${d.shareChg>=0?'badge-green':'badge-red'}">${fmtPP(d.shareChg)}</span></td>
    </tr>
  `).join('') + `
    <tr class="total-row"><td></td><td><b>TOTAL</b></td>
      <td class="align-right"><b>${fmt(totVol)}</b></td><td class="align-right"><b>${fmt(totYoy)}</b></td>
      <td class="align-right ${totGrowth>=0?'positive':'negative'}"><b>${totGrowth>=0?'+':''}${totGrowth.toFixed(1)}%</b></td>
      <td class="align-right"><b>100.0%</b></td><td class="align-right"></td>
    </tr>`;
}

// ============================================
// COMPANY DEEP-DIVE
// ============================================
function renderCompanyView() {
  const c = currentCompany, sub = currentSubseg;
  if (!c) return;

  const indQ = getIndustryVols(sub);
  const cvQ = getCompanyVols(c,sub);
  const curVol = periodVol(cvQ), yoyV = yoyPeriodVol(cvQ);
  const curInd = periodVol(indQ), yoyInd = yoyPeriodVol(indQ);
  const growth = yoyV>0?((curVol/yoyV-1)*100):0;
  const shareCur = curInd>0?(curVol/curInd*100):0;
  const shareYoy = yoyInd>0?(yoyV/yoyInd*100):0;
  const shareChg = shareCur-shareYoy;

  document.getElementById('kpi-company').innerHTML = `
    <div class="kpi"><div class="kpi-label">Volume (${periodLabel()})</div><div class="kpi-value">${fmt(curVol)}</div></div>
    <div class="kpi"><div class="kpi-label">Market Share</div><div class="kpi-value">${shareCur.toFixed(1)}%</div><div class="kpi-sub ${shareChg>=0?'positive':'negative'}">${fmtPP(shareChg)} YoY</div></div>
    <div class="kpi"><div class="kpi-label">YoY Volume Growth</div><div class="kpi-value ${growth>=0?'positive':'negative'}">${growth>=0?'+':''}${growth.toFixed(1)}%</div></div>
    <div class="kpi"><div class="kpi-label">Industry Rank</div><div class="kpi-value">#${topCompanies(50,sub,null).indexOf(c)+1}</div></div>
  `;

  const vm = getViewMode();
  document.getElementById('title-co-vol').textContent = 'Volume Trend (' + (vm==='annual'?'Annual':'Quarterly') + ')';
  const geoLabel = currentGeo === 'zone' ? 'Zone' : 'State';
  document.getElementById('title-co-states').textContent = 'Top ' + geoLabel + 's by Volume (' + periodLabel() + ')';
  document.getElementById('title-co-share-chg').textContent = 'Mkt Share Change by ' + geoLabel + ' (YoY, pp) - ' + periodLabel();
  document.getElementById('title-co-contrib').textContent = 'Sales Contribution Trend by ' + geoLabel + ' (%)';
  document.getElementById('title-co-share-ts').textContent = 'Market Share Time-Series by ' + geoLabel + ' (%)';
  document.getElementById('title-co-geo-details').textContent = geoLabel + '-wise Details';

  // Volume trend
  const tsC = tsVols(cvQ), tsI = tsVols(indQ);
  Plotly.newPlot('chart-co-vol',[
    {x:tsDates(),y:tsC,type:'scatter',mode:'lines+markers',name:c+' Volume',line:{color:getColor(c,0),width:2.5},marker:{size:4},hovertemplate:c+': %{y:,.0f}<extra></extra>'},
    {x:tsDates(),y:tsI,type:'scatter',mode:'lines',name:'Industry',yaxis:'y2',line:{color:'#94a3b8',width:1.5,dash:'dot'},hovertemplate:'Industry: %{y:,.0f}<extra></extra>'}
  ],{...PLOTLY_LAYOUT,margin:{l:80,r:80,t:10,b:75},yaxis:{title:{text:c,standoff:12},showgrid:true,gridcolor:'#f0f0f0',automargin:true},yaxis2:{title:{text:'Industry',standoff:12},overlaying:'y',side:'right',showgrid:false,automargin:true},legend:{orientation:'h',y:-0.22,x:0.5,xanchor:'center',font:{size:10}}},PLOTLY_CONFIG);
  addChartCopyBtn('chart-co-vol');

  // Market share trend
  const shareTS = computeShare(tsC, tsI);
  plotLines('chart-co-share',[{
    x:tsDates(),y:shareTS,type:'scatter',mode:'lines+markers',name:'Market Share',
    line:{color:getColor(c,0),width:2.5},marker:{size:4},
    fill:'tozeroy',fillcolor:getColor(c,0)+'15',
    hovertemplate:'Share: %{y:.1f}%<extra></extra>'
  }],'Market Share (%)',true);

  // YoY growth trend
  const coYoY = yoyGrowthSeries(cvQ);
  const indYoY = yoyGrowthSeries(indQ);
  const dates = tsDates();
  plotYoYGrowth('chart-co-yoy',[
    {x:dates,y:vm==='annual'?coYoY:coYoY,type:'scatter',mode:'lines+markers',name:c,line:{color:getColor(c,0),width:2.5},marker:{size:4},connectgaps:false,hovertemplate:c+': %{y:.1f}%<extra></extra>'},
    {x:dates,y:vm==='annual'?indYoY:indYoY,type:'scatter',mode:'lines',name:'Industry',line:{color:'#94a3b8',width:1.5,dash:'dot'},connectgaps:false,hovertemplate:'Industry: %{y:.1f}%<extra></extra>'}
  ]);

  // Subsegment mix section for company (PV: Cars/UVs, 2W: Motorcycle/Scooters)
  var actualSubsegsCo = segSubsegs.filter(function(s){return s!=='All';});
  var mixElCo = document.getElementById('subseg-mix-company');
  if (actualSubsegsCo.length >= 2 && sub === 'All') {
    mixElCo.style.display = 'block';
    var segNameCo = currentSegment;
    var subColorsCo = ['#2563eb','#f59e0b','#10b981','#ef4444','#8b5cf6'];
    var subVolsCo = actualSubsegsCo.map(function(s){return tsVols(getCompanyVols(c,s));});
    var mixTotalsCo = subVolsCo[0].map(function(_,t){return subVolsCo.reduce(function(sum,sv){return sum+sv[t];},0);});
    var subPctsCo = subVolsCo.map(function(sv){return sv.map(function(v,t){return mixTotalsCo[t]>0?v/mixTotalsCo[t]*100:0;});});
    var datesCo = tsDates(), tsLenCo = datesCo.length;

    document.getElementById('title-co-subseg-mix').textContent = c + ' - ' + segNameCo + ' Subsegment Mix (%)';
    var mixTracesCo = actualSubsegsCo.map(function(s,i){return {
      x:datesCo, y:subPctsCo[i], customdata:subVolsCo[i], type:'bar', name:s,
      marker:{color:subColorsCo[i%subColorsCo.length]},
      hovertemplate:s+': %{y:.1f}%<br>Volume: %{customdata:,.0f}<extra></extra>'
    };});
    Plotly.newPlot('chart-co-subseg-mix',mixTracesCo,{
      ...PLOTLY_LAYOUT, barmode:'stack',
      margin:{l:60,r:25,t:10,b:75},
      yaxis:{...PLOTLY_LAYOUT.yaxis, title:{text:'% Mix',standoff:10}, range:[0,100.5]},
      legend:{orientation:'h',y:-0.22,x:0.5,xanchor:'center',font:{size:10}}
    },PLOTLY_CONFIG);
    addChartCopyBtn('chart-co-subseg-mix');

    document.getElementById('title-co-subseg-vol').textContent = c + ' - ' + segNameCo + ' Subsegment Volumes';
    var volHeadCo = '<tr><th>Subsegment</th>' + datesCo.map(function(d){return '<th class="align-right">'+d+'</th>';}).join('') + '</tr>';
    var volBodyCo = '';
    actualSubsegsCo.forEach(function(s,i){
      volBodyCo += '<tr><td><b>'+s+'</b></td>';
      for(var t=0;t<tsLenCo;t++) volBodyCo += '<td class="align-right">'+fmt(subVolsCo[i][t])+'</td>';
      volBodyCo += '</tr>';
    });
    volBodyCo += '<tr class="total-row"><td><b>Total '+segNameCo+'</b></td>';
    for(var t=0;t<tsLenCo;t++) volBodyCo += '<td class="align-right"><b>'+fmt(mixTotalsCo[t])+'</b></td>';
    volBodyCo += '</tr>';
    actualSubsegsCo.forEach(function(s,i){
      volBodyCo += '<tr style="color:#6b7280;font-size:12px"><td><i>'+s+' %</i></td>';
      for(var t=0;t<tsLenCo;t++) volBodyCo += '<td class="align-right"><i>'+subPctsCo[i][t].toFixed(1)+'%</i></td>';
      volBodyCo += '</tr>';
    });
    document.querySelector('#table-co-subseg-vol thead').innerHTML = volHeadCo;
    document.querySelector('#table-co-subseg-vol tbody').innerHTML = volBodyCo;

    document.getElementById('title-co-subseg-yoy').textContent = c + ' - ' + segNameCo + ' Subsegment YoY Growth (%)';
    var yoyHeadCo = '<tr><th>Subsegment</th>' + datesCo.map(function(d){return '<th class="align-right">'+d+'</th>';}).join('') + '</tr>';
    var yoyBodyCo = '';
    actualSubsegsCo.forEach(function(s){
      var yoySC = yoyGrowthSeries(getCompanyVols(c,s));
      yoyBodyCo += '<tr><td><b>'+s+'</b></td>';
      for(var t=0;t<tsLenCo;t++){
        var gC=yoySC[t];
        if(gC===null||gC===undefined) yoyBodyCo+='<td class="align-right">-</td>';
        else yoyBodyCo+='<td class="align-right '+(gC>=0?'positive':'negative')+'">'+(gC>=0?'+':'')+gC.toFixed(1)+'%</td>';
      }
      yoyBodyCo += '</tr>';
    });
    var overallYoYCo = yoyGrowthSeries(cvQ);
    yoyBodyCo += '<tr class="total-row"><td><b>Overall '+segNameCo+'</b></td>';
    for(var t=0;t<tsLenCo;t++){
      var g2C=overallYoYCo[t];
      if(g2C===null||g2C===undefined) yoyBodyCo+='<td class="align-right">-</td>';
      else yoyBodyCo+='<td class="align-right '+(g2C>=0?'positive':'negative')+'"><b>'+(g2C>=0?'+':'')+g2C.toFixed(1)+'%</b></td>';
    }
    yoyBodyCo += '</tr>';
    document.querySelector('#table-co-subseg-yoy thead').innerHTML = yoyHeadCo;
    document.querySelector('#table-co-subseg-yoy tbody').innerHTML = yoyBodyCo;
  } else {
    mixElCo.style.display = 'none';
  }

  // Geo-level volume (selected period) — works for both state and zone
  const geoList = currentGeo === 'zone' ? segZones : segStates;
  const getGeoIndVols = currentGeo === 'zone'
    ? (g,sub2) => getZoneIndustryVols(g,sub2)
    : (g,sub2) => getStateIndustryVols(g,sub2);
  const getGeoCoVols = currentGeo === 'zone'
    ? (g,co2,sub2) => getZoneCompanyVols(g,co2,sub2)
    : (g,co2,sub2) => getStateCompanyVols(g,co2,sub2);

  const geoVols = {};
  const pQs = periodQIdxs();
  for (const g of geoList) {
    const cv = getGeoCoVols(g,c,sub);
    const vol = pQs.reduce((s,qi)=>s+(cv[qi]||0),0);
    if (vol > 0) geoVols[g] = vol;
  }
  const sortedGeos = Object.entries(geoVols).sort((a,b)=>b[1]-a[1]);
  const topN = sortedGeos.slice(0,15);
  plotHBar('chart-co-states',
    topN.map(e=>e[0]).reverse(), topN.map(e=>e[1]).reverse(),
    topN.map(()=>getColor(c,0)).reverse()
  );

  // Share change by geo
  const yoyQs = yoyPeriodQIdxs();
  const shareChgByGeo = [];
  for (const [g] of sortedGeos) {
    const gIndQ = getGeoIndVols(g,sub);
    const gCoQ = getGeoCoVols(g,c,sub);
    const curG = pQs.reduce((s,qi)=>s+gCoQ[qi],0);
    const curGInd = pQs.reduce((s,qi)=>s+gIndQ[qi],0);
    const yoyG = yoyQs.reduce((s,qi)=>s+(gCoQ[qi]||0),0);
    const yoyGInd = yoyQs.reduce((s,qi)=>s+(gIndQ[qi]||0),0);
    const sCur = curGInd>0?curG/curGInd*100:0;
    const sYoy = yoyGInd>0?yoyG/yoyGInd*100:0;
    if (sCur>0||sYoy>0) shareChgByGeo.push({geo:g,chg:sCur-sYoy,curShare:sCur});
  }
  shareChgByGeo.sort((a,b)=>b.chg-a.chg);
  const displayChg = shareChgByGeo.slice(0,18);
  plotHBarDiverging('chart-co-share-chg',
    displayChg.map(e=>e.geo).reverse(),
    displayChg.map(e=>parseFloat(e.chg.toFixed(1))).reverse(),
    displayChg.map(e=>e.chg>=0?'#059669':'#dc2626').reverse()
  );

  // --- Contribution Time-Series Chart ---
  const top10Geos = sortedGeos.slice(0,10).map(e=>e[0]);
  const tsLen2 = tsLen();
  const contribTraces = [];
  const coTotalTS = tsVols(cvQ); // company total volumes time-series
  top10Geos.forEach((g,gi) => {
    const gCoQ = getGeoCoVols(g,c,sub);
    const gTS = tsVols(gCoQ);
    const contribTS = gTS.map((v,t) => coTotalTS[t]>0 ? v/coTotalTS[t]*100 : 0);
    contribTraces.push({
      x:tsDates(), y:contribTS, type:'scatter', mode:'lines+markers',
      name:g, line:{color:PALETTE[gi%PALETTE.length],width:2}, marker:{size:3},
      hovertemplate:g+': %{y:.1f}%<extra></extra>'
    });
  });
  // Others contribution line
  if (sortedGeos.length > 10) {
    const othersContrib = tsDates().map((_,t) => {
      const topSum = top10Geos.reduce((s,g) => {
        const gCoQ = getGeoCoVols(g,c,sub);
        return s + tsVols(gCoQ)[t];
      }, 0);
      return coTotalTS[t]>0 ? (coTotalTS[t]-topSum)/coTotalTS[t]*100 : 0;
    });
    contribTraces.push({
      x:tsDates(), y:othersContrib, type:'scatter', mode:'lines',
      name:'Others', line:{color:'#d1d5db',width:1.5,dash:'dot'},
      hovertemplate:'Others: %{y:.1f}%<extra></extra>'
    });
  }
  plotLines('chart-co-contrib', contribTraces, 'Contribution (%)', true);

  // --- Market Share Time-Series Table ---
  const tsLabels2 = tsDates();
  const shareTSHead = document.querySelector('#table-co-share-ts thead');
  const shareTSBody = document.querySelector('#table-co-share-ts tbody');
  let shareHdr = '<tr><th>' + geoLabel + '</th>';
  for (let t = 0; t < tsLen2; t++) shareHdr += '<th class="align-right">' + tsLabels2[t] + '</th>';
  shareHdr += '</tr>';
  shareTSHead.innerHTML = shareHdr;

  let shareRows = '';
  for (let gi = 0; gi < Math.min(sortedGeos.length, 10); gi++) {
    const g = sortedGeos[gi][0];
    const gCoQ = getGeoCoVols(g,c,sub);
    const gIndQ = getGeoIndVols(g,sub);
    const gCoTS = tsVols(gCoQ);
    const gIndTS = tsVols(gIndQ);
    shareRows += '<tr><td><b>' + g + '</b></td>';
    for (let t = 0; t < tsLen2; t++) {
      const sh = gIndTS[t]>0 ? gCoTS[t]/gIndTS[t]*100 : 0;
      const bg = sh > 20 ? 'rgba(5,150,105,0.15)' : sh > 10 ? 'rgba(5,150,105,0.07)' : '';
      shareRows += '<td class="align-right" style="background:' + bg + '">' + sh.toFixed(1) + '%</td>';
    }
    shareRows += '</tr>';
  }
  // Others row
  if (sortedGeos.length > 10) {
    const othersGeos = sortedGeos.slice(10).map(e=>e[0]);
    shareRows += '<tr><td><b>Others</b></td>';
    for (let t = 0; t < tsLen2; t++) {
      let oCo = 0, oInd = 0;
      for (const g of othersGeos) {
        oCo += tsVols(getGeoCoVols(g,c,sub))[t];
        oInd += tsVols(getGeoIndVols(g,sub))[t];
      }
      const sh = oInd>0 ? oCo/oInd*100 : 0;
      shareRows += '<td class="align-right">' + sh.toFixed(1) + '%</td>';
    }
    shareRows += '</tr>';
  }
  // Overall national market share row
  const natCoTS = tsVols(cvQ);
  const natIndTS = tsVols(indQ);
  shareRows += '<tr class="total-row" style="border-top:2px solid #94a3b8"><td><b>Overall (National)</b></td>';
  for (let t = 0; t < tsLen2; t++) {
    const natSh = natIndTS[t]>0 ? natCoTS[t]/natIndTS[t]*100 : 0;
    shareRows += '<td class="align-right"><b>' + natSh.toFixed(1) + '%</b></td>';
  }
  shareRows += '</tr>';
  shareTSBody.innerHTML = shareRows;

  // Geo details table
  const totalCoVol = curVol;
  const geoDetails = sortedGeos.map(([g,vol]) => {
    const gIndQ = getGeoIndVols(g,sub);
    const gCoQ = getGeoCoVols(g,c,sub);
    const sVol = pQs.reduce((s,qi)=>s+(gCoQ[qi]||0),0);
    const sYoyVol = yoyQs.reduce((s,qi)=>s+(gCoQ[qi]||0),0);
    const sGrowth = sYoyVol>0?((sVol/sYoyVol-1)*100):0;
    const sIndCur = pQs.reduce((s,qi)=>s+(gIndQ[qi]||0),0);
    const sIndYoy = yoyQs.reduce((s,qi)=>s+(gIndQ[qi]||0),0);
    const sShare = sIndCur>0?sVol/sIndCur*100:0;
    const sShareYoy = sIndYoy>0?sYoyVol/sIndYoy*100:0;
    let zone = '';
    if (currentGeo === 'state') {
      const r = filterRows(c,g,sub);
      zone = r.length>0 ? ROWS[r[0]][2] : '';
    }
    return {geo:g,zone,vol:sVol,yoyVol:sYoyVol,growth:sGrowth,share:sShare,shareChg:sShare-sShareYoy,contrib:totalCoVol>0?sVol/totalCoVol*100:0};
  }).filter(d=>d.vol>0);

  // Totals
  const totVol2=geoDetails.reduce((s,d)=>s+d.vol,0);
  const totYoy2=geoDetails.reduce((s,d)=>s+d.yoyVol,0);
  const totGrowth2=totYoy2>0?((totVol2/totYoy2-1)*100):0;

  const thead = document.querySelector('#table-co-states thead');
  const zoneCol = currentGeo === 'state' ? '<th>Zone</th>' : '';
  thead.innerHTML = '<tr><th>#</th><th>' + geoLabel + '</th>' + zoneCol + '<th class="align-right">Volume</th><th class="align-right">YoY Vol</th><th class="align-right">YoY Growth</th><th class="align-right">Mkt Share</th><th class="align-right">Share Chg</th><th class="align-right">Contribution</th></tr>';
  const tbody2 = document.querySelector('#table-co-states tbody');
  tbody2.innerHTML = geoDetails.map((d,i) => {
    const zoneTd = currentGeo === 'state' ? '<td>' + d.zone + '</td>' : '';
    const onclick = currentGeo === 'state' ? `onclick="drillToState('${esc(d.geo)}','${esc(d.zone)}')"` : '';
    return `<tr class="${currentGeo==='state'?'clickable':''}" ${onclick}>
      <td>${i+1}</td><td><b>${d.geo}</b></td>${zoneTd}
      <td class="align-right">${fmt(d.vol)}</td><td class="align-right">${fmt(d.yoyVol)}</td>
      <td class="align-right ${d.growth>=0?'positive':'negative'}">${d.growth>=0?'+':''}${d.growth.toFixed(1)}%</td>
      <td class="align-right">${d.share.toFixed(1)}%</td>
      <td class="align-right"><span class="badge ${d.shareChg>=0?'badge-green':'badge-red'}">${fmtPP(d.shareChg)}</span></td>
      <td class="align-right">${d.contrib.toFixed(1)}%</td>
    </tr>`;
  }).join('') + `
    <tr class="total-row"><td></td><td><b>TOTAL</b></td>${currentGeo==='state'?'<td></td>':''}
      <td class="align-right"><b>${fmt(totVol2)}</b></td><td class="align-right"><b>${fmt(totYoy2)}</b></td>
      <td class="align-right ${totGrowth2>=0?'positive':'negative'}"><b>${totGrowth2>=0?'+':''}${totGrowth2.toFixed(1)}%</b></td>
      <td class="align-right"><b>${shareCur.toFixed(1)}%</b></td><td class="align-right"><span class="badge ${shareChg>=0?'badge-green':'badge-red'}">${fmtPP(shareChg)}</span></td>
      <td class="align-right"><b>100.0%</b></td>
    </tr>`;
}

// ============================================
// STATE DEEP-DIVE
// ============================================
function renderStateView() {
  const st = currentState, sub = currentSubseg;
  if (!st) return;

  const stIndQ = getStateIndustryVols(st,sub);
  const curVol = periodVol(stIndQ);
  const yoyV = yoyPeriodVol(stIndQ);
  const growth = yoyV>0?((curVol/yoyV-1)*100):0;
  const natVol = periodVol(getIndustryVols(sub));

  const top = topCompanies(8,sub,st);
  const topShares = top.map(c => {
    const cv = getStateCompanyVols(st,c,sub);
    const v = periodVol(cv);
    return {company:c, share:curVol>0?v/curVol*100:0};
  });

  const vm = getViewMode();
  document.getElementById('title-st-vol').textContent = 'Industry Volume Trend (' + (vm==='annual'?'Annual':'Quarterly') + ')';
  document.getElementById('title-st-pie').textContent = 'Market Share Breakdown (' + periodLabel() + ')';

  document.getElementById('kpi-state').innerHTML = `
    <div class="kpi"><div class="kpi-label">State Volume (${periodLabel()})</div><div class="kpi-value">${fmt(curVol)}</div><div class="kpi-sub ${growth>=0?'positive':'negative'}">YoY: ${growth>=0?'+':''}${growth.toFixed(1)}%</div></div>
    <div class="kpi"><div class="kpi-label">% of National Volume</div><div class="kpi-value">${(natVol>0?(curVol/natVol*100):0).toFixed(1)}%</div></div>
    <div class="kpi"><div class="kpi-label">#1 Company</div><div class="kpi-value" style="font-size:18px">${topShares[0]?.company||'-'}</div><div class="kpi-sub neutral">${topShares[0]?.share.toFixed(1)||0}% share</div></div>
    <div class="kpi"><div class="kpi-label">#2 Company</div><div class="kpi-value" style="font-size:18px">${topShares[1]?.company||'-'}</div><div class="kpi-sub neutral">${topShares[1]?.share.toFixed(1)||0}% share</div></div>
  `;

  // Volume trend
  const tsV = tsVols(stIndQ);
  plotLines('chart-st-vol',[{
    x:tsDates(),y:tsV,type:'scatter',mode:'lines+markers',name:'Volume',
    line:{color:'#2563eb',width:2.5},marker:{size:4},
    fill:'tozeroy',fillcolor:'rgba(37,99,235,0.08)',
    hovertemplate:'%{x}: %{y:,.0f}<extra></extra>'
  }],'Volume (units)',false);

  // Market share donut (selected period)
  const pQs = periodQIdxs();
  const allComps = topCompanies(20,sub,st);
  const pieLabels=[],pieValues=[],pieColors=[];
  let othersVol=0;
  allComps.forEach((c,ci) => {
    const cv = getStateCompanyVols(st,c,sub);
    const v = pQs.reduce((s,qi)=>s+(cv[qi]||0),0);
    if (ci<8 && v>0){pieLabels.push(c);pieValues.push(v);pieColors.push(getColor(c,ci));}
    else othersVol+=v;
  });
  if (othersVol>0){pieLabels.push('Others');pieValues.push(othersVol);pieColors.push('#d1d5db');}
  plotDonut('chart-st-pie',pieLabels,pieValues,pieColors);

  // Market share trend
  const top10 = topCompanies(10,sub,st);
  const shareTraces = top10.map((c,ci) => {
    const cv = getStateCompanyVols(st,c,sub);
    const share = computeShare(tsVols(cv), tsV);
    return {x:tsDates(),y:share,type:'scatter',mode:'lines',name:c,line:{color:getColor(c,ci),width:2},
      hovertemplate:c+': %{y:.1f}%<extra></extra>'};
  });
  plotLines('chart-st-share',shareTraces,'Market Share (%)',true);

  // YoY growth trend
  const stYoY = yoyGrowthSeries(stIndQ);
  const natYoY = yoyGrowthSeries(getIndustryVols(sub));
  const top3 = topCompanies(3,sub,st);
  const yoyTraces = [
    {x:tsDates(),y:stYoY,type:'scatter',mode:'lines+markers',name:st+' Industry',line:{color:'#2563eb',width:2.5},marker:{size:4},connectgaps:false,hovertemplate:st+': %{y:.1f}%<extra></extra>'},
    {x:tsDates(),y:natYoY,type:'scatter',mode:'lines',name:'National Industry',line:{color:'#94a3b8',width:1.5,dash:'dot'},connectgaps:false,hovertemplate:'National: %{y:.1f}%<extra></extra>'}
  ];
  top3.forEach((c,ci) => {
    const yoy = yoyGrowthSeries(getStateCompanyVols(st,c,sub));
    yoyTraces.push({x:tsDates(),y:yoy,type:'scatter',mode:'lines',name:c,line:{color:getColor(c,ci+2),width:1.5},connectgaps:false,hovertemplate:c+': %{y:.1f}%<extra></extra>'});
  });
  plotYoYGrowth('chart-st-yoy',yoyTraces);

  // Subsegment mix section (PV: Cars/UVs, 2W: Motorcycle/Scooters)
  var actualSubsegs = segSubsegs.filter(function(s){return s!=='All';});
  var mixEl = document.getElementById('subseg-mix-state');
  if (actualSubsegs.length >= 2 && sub === 'All') {
    mixEl.style.display = 'block';
    var segName = currentSegment;
    var subColors = ['#2563eb','#f59e0b','#10b981','#ef4444','#8b5cf6'];
    var subVols = actualSubsegs.map(function(s){return tsVols(getStateIndustryVols(st,s));});
    var mixTotals = subVols[0].map(function(_,t){return subVols.reduce(function(sum,sv){return sum+sv[t];},0);});
    var subPcts = subVols.map(function(sv){return sv.map(function(v,t){return mixTotals[t]>0?v/mixTotals[t]*100:0;});});
    var dates = tsDates(), tsLen = dates.length;

    /* 100% stacked bar chart */
    document.getElementById('title-st-subseg-mix').textContent = segName + ' Subsegment Mix (%)';
    var mixTraces = actualSubsegs.map(function(s,i){return {
      x:dates, y:subPcts[i], customdata:subVols[i], type:'bar', name:s,
      marker:{color:subColors[i%subColors.length]},
      hovertemplate:s+': %{y:.1f}%<br>Volume: %{customdata:,.0f}<extra></extra>'
    };});
    Plotly.newPlot('chart-st-subseg-mix',mixTraces,{
      ...PLOTLY_LAYOUT, barmode:'stack',
      margin:{l:60,r:25,t:10,b:75},
      yaxis:{...PLOTLY_LAYOUT.yaxis, title:{text:'% Mix',standoff:10}, range:[0,100.5]},
      legend:{orientation:'h',y:-0.22,x:0.5,xanchor:'center',font:{size:10}}
    },PLOTLY_CONFIG);
    addChartCopyBtn('chart-st-subseg-mix');

    /* Absolute sales table */
    document.getElementById('title-st-subseg-vol').textContent = segName + ' Subsegment Volumes';
    var volHead = '<tr><th>Subsegment</th>' + dates.map(function(d){return '<th class="align-right">'+d+'</th>';}).join('') + '</tr>';
    var volBody = '';
    actualSubsegs.forEach(function(s,i){
      volBody += '<tr><td><b>'+s+'</b></td>';
      for(var t=0;t<tsLen;t++) volBody += '<td class="align-right">'+fmt(subVols[i][t])+'</td>';
      volBody += '</tr>';
    });
    volBody += '<tr class="total-row"><td><b>Total '+segName+'</b></td>';
    for(var t=0;t<tsLen;t++) volBody += '<td class="align-right"><b>'+fmt(mixTotals[t])+'</b></td>';
    volBody += '</tr>';
    actualSubsegs.forEach(function(s,i){
      volBody += '<tr style="color:#6b7280;font-size:12px"><td><i>'+s+' %</i></td>';
      for(var t=0;t<tsLen;t++) volBody += '<td class="align-right"><i>'+subPcts[i][t].toFixed(1)+'%</i></td>';
      volBody += '</tr>';
    });
    document.querySelector('#table-st-subseg-vol thead').innerHTML = volHead;
    document.querySelector('#table-st-subseg-vol tbody').innerHTML = volBody;

    /* YoY growth table */
    document.getElementById('title-st-subseg-yoy').textContent = segName + ' Subsegment YoY Growth (%)';
    var yoyHead2 = '<tr><th>Subsegment</th>' + dates.map(function(d){return '<th class="align-right">'+d+'</th>';}).join('') + '</tr>';
    var yoyBody2 = '';
    actualSubsegs.forEach(function(s){
      var yoy2 = yoyGrowthSeries(getStateIndustryVols(st,s));
      yoyBody2 += '<tr><td><b>'+s+'</b></td>';
      for(var t=0;t<tsLen;t++){
        var g=yoy2[t];
        if(g===null||g===undefined) yoyBody2+='<td class="align-right">-</td>';
        else yoyBody2+='<td class="align-right '+(g>=0?'positive':'negative')+'">'+(g>=0?'+':'')+g.toFixed(1)+'%</td>';
      }
      yoyBody2 += '</tr>';
    });
    var overallYoY = yoyGrowthSeries(stIndQ);
    yoyBody2 += '<tr class="total-row"><td><b>Overall '+segName+'</b></td>';
    for(var t=0;t<tsLen;t++){
      var g2=overallYoY[t];
      if(g2===null||g2===undefined) yoyBody2+='<td class="align-right">-</td>';
      else yoyBody2+='<td class="align-right '+(g2>=0?'positive':'negative')+'"><b>'+(g2>=0?'+':'')+g2.toFixed(1)+'%</b></td>';
    }
    yoyBody2 += '</tr>';
    document.querySelector('#table-st-subseg-yoy thead').innerHTML = yoyHead2;
    document.querySelector('#table-st-subseg-yoy tbody').innerHTML = yoyBody2;
  } else {
    mixEl.style.display = 'none';
  }

  // Company table
  const yoyQs = yoyPeriodQIdxs();
  const comps = topCompanies(50,sub,st);
  const tableData = comps.map(c => {
    const cv = getStateCompanyVols(st,c,sub);
    const v = pQs.reduce((s,qi)=>s+(cv[qi]||0),0);
    const yv = yoyQs.reduce((s,qi)=>s+(cv[qi]||0),0);
    const g = yv>0?((v/yv-1)*100):0;
    const s = curVol>0?v/curVol*100:0;
    const stIndYoy = yoyQs.reduce((sum,qi)=>sum+(stIndQ[qi]||0),0);
    const sy = stIndYoy>0?yv/stIndYoy*100:0;
    return {company:c,vol:v,yoyVol:yv,growth:g,share:s,shareChg:s-sy};
  }).filter(d=>d.vol>0);
  tableData.sort((a,b)=>b.vol-a.vol);

  const totVol=tableData.reduce((s,d)=>s+d.vol,0);
  const totYoy=tableData.reduce((s,d)=>s+d.yoyVol,0);
  const totG=totYoy>0?((totVol/totYoy-1)*100):0;

  const thead = document.querySelector('#table-st-companies thead');
  thead.innerHTML = '<tr><th>#</th><th>Company</th><th class="align-right">Volume</th><th class="align-right">YoY Vol</th><th class="align-right">YoY Growth</th><th class="align-right">Mkt Share</th><th class="align-right">Share Chg</th></tr>';
  const tbody = document.querySelector('#table-st-companies tbody');
  tbody.innerHTML = tableData.map((d,i) => `
    <tr class="clickable" onclick="drillToCompany('${esc(d.company)}')">
      <td>${i+1}</td><td><b>${d.company}</b></td>
      <td class="align-right">${fmt(d.vol)}</td><td class="align-right">${fmt(d.yoyVol)}</td>
      <td class="align-right ${d.growth>=0?'positive':'negative'}">${d.growth>=0?'+':''}${d.growth.toFixed(1)}%</td>
      <td class="align-right">${d.share.toFixed(1)}%</td>
      <td class="align-right"><span class="badge ${d.shareChg>=0?'badge-green':'badge-red'}">${fmtPP(d.shareChg)}</span></td>
    </tr>`).join('') + `
    <tr class="total-row"><td></td><td><b>TOTAL</b></td>
      <td class="align-right"><b>${fmt(totVol)}</b></td><td class="align-right"><b>${fmt(totYoy)}</b></td>
      <td class="align-right ${totG>=0?'positive':'negative'}"><b>${totG>=0?'+':''}${totG.toFixed(1)}%</b></td>
      <td class="align-right"><b>100.0%</b></td><td class="align-right"></td>
    </tr>`;
}

// ============================================
// ZONE DEEP-DIVE
// ============================================
function renderZoneView() {
  const zone = currentZoneTab, sub = currentSubseg;
  if (!zone) return;

  const znIndQ = getZoneIndustryVols(zone, sub);
  const curVol = periodVol(znIndQ);
  const yoyV = yoyPeriodVol(znIndQ);
  const growth = yoyV > 0 ? ((curVol/yoyV-1)*100) : 0;
  const natVol = periodVol(getIndustryVols(sub));

  const top = topCompanies(8, sub, null, zone);
  const topShares = top.map(c => {
    const cv = getZoneCompanyVols(zone, c, sub);
    const v = periodVol(cv);
    return {company:c, share:curVol>0?v/curVol*100:0};
  });

  const vm = getViewMode();
  document.getElementById('title-zn-vol').textContent = 'Zone Volume Trend (' + (vm==='annual'?'Annual':'Quarterly') + ')';
  document.getElementById('title-zn-pie').textContent = 'Market Share Breakdown (' + periodLabel() + ')';
  document.getElementById('title-zn-states').textContent = 'States by Volume (' + periodLabel() + ')';
  document.getElementById('title-zn-state-contrib').textContent = 'State Contribution Trend (%)';
  document.getElementById('title-zn-state-table').textContent = 'State-wise Details (' + periodLabel() + ')';
  document.getElementById('title-zn-comp-table').textContent = 'Company Rankings in ' + zone + ' (' + periodLabel() + ')';

  // KPIs
  document.getElementById('kpi-zone').innerHTML = `
    <div class="kpi"><div class="kpi-label">Zone Volume (${periodLabel()})</div><div class="kpi-value">${fmt(curVol)}</div><div class="kpi-sub ${growth>=0?'positive':'negative'}">YoY: ${growth>=0?'+':''}${growth.toFixed(1)}%</div></div>
    <div class="kpi"><div class="kpi-label">% of National Volume</div><div class="kpi-value">${(natVol>0?(curVol/natVol*100):0).toFixed(1)}%</div></div>
    <div class="kpi"><div class="kpi-label">#1 Company</div><div class="kpi-value" style="font-size:18px">${topShares[0]?.company||'-'}</div><div class="kpi-sub neutral">${topShares[0]?.share.toFixed(1)||0}% share</div></div>
    <div class="kpi"><div class="kpi-label">#2 Company</div><div class="kpi-value" style="font-size:18px">${topShares[1]?.company||'-'}</div><div class="kpi-sub neutral">${topShares[1]?.share.toFixed(1)||0}% share</div></div>
  `;

  // Volume trend
  const tsV = tsVols(znIndQ);
  plotLines('chart-zn-vol',[{
    x:tsDates(),y:tsV,type:'scatter',mode:'lines+markers',name:'Volume',
    line:{color:'#2563eb',width:2.5},marker:{size:4},
    fill:'tozeroy',fillcolor:'rgba(37,99,235,0.08)',
    hovertemplate:'%{x}: %{y:,.0f}<extra></extra>'
  }],'Volume (units)',false);

  // Market share donut (selected period)
  const pQs = periodQIdxs();
  const allComps = topCompanies(20, sub, null, zone);
  const pieLabels=[],pieValues=[],pieColors=[];
  let othersVol=0;
  allComps.forEach((c,ci) => {
    const cv = getZoneCompanyVols(zone, c, sub);
    const v = pQs.reduce((s,qi)=>s+(cv[qi]||0),0);
    if (ci<8 && v>0){pieLabels.push(c);pieValues.push(v);pieColors.push(getColor(c,ci));}
    else othersVol+=v;
  });
  if (othersVol>0){pieLabels.push('Others');pieValues.push(othersVol);pieColors.push('#d1d5db');}
  plotDonut('chart-zn-pie',pieLabels,pieValues,pieColors);

  // Market share trend - top 10 companies in zone
  const top10 = topCompanies(10, sub, null, zone);
  const shareTraces = top10.map((c,ci) => {
    const cv = getZoneCompanyVols(zone, c, sub);
    const share = computeShare(tsVols(cv), tsV);
    return {x:tsDates(),y:share,type:'scatter',mode:'lines',name:c,line:{color:getColor(c,ci),width:2},
      hovertemplate:c+': %{y:.1f}%<extra></extra>'};
  });
  plotLines('chart-zn-share',shareTraces,'Market Share (%)',true);

  // YoY growth trend - zone vs national + top 3 companies
  const znYoY = yoyGrowthSeries(znIndQ);
  const natYoY = yoyGrowthSeries(getIndustryVols(sub));
  const top3 = topCompanies(3, sub, null, zone);
  const yoyTraces = [
    {x:tsDates(),y:znYoY,type:'scatter',mode:'lines+markers',name:zone+' Industry',line:{color:'#2563eb',width:2.5},marker:{size:4},connectgaps:false,hovertemplate:zone+': %{y:.1f}%<extra></extra>'},
    {x:tsDates(),y:natYoY,type:'scatter',mode:'lines',name:'National Industry',line:{color:'#94a3b8',width:1.5,dash:'dot'},connectgaps:false,hovertemplate:'National: %{y:.1f}%<extra></extra>'}
  ];
  top3.forEach((c,ci) => {
    const yoy = yoyGrowthSeries(getZoneCompanyVols(zone, c, sub));
    yoyTraces.push({x:tsDates(),y:yoy,type:'scatter',mode:'lines',name:c,line:{color:getColor(c,ci+2),width:1.5},connectgaps:false,hovertemplate:c+': %{y:.1f}%<extra></extra>'});
  });
  plotYoYGrowth('chart-zn-yoy',yoyTraces);

  // Subsegment mix section for zone (PV: Cars/UVs, 2W: Motorcycle/Scooters)
  var actualSubsegsZn = segSubsegs.filter(function(s){return s!=='All';});
  var mixElZn = document.getElementById('subseg-mix-zone');
  if (actualSubsegsZn.length >= 2 && sub === 'All') {
    mixElZn.style.display = 'block';
    var segNameZn = currentSegment;
    var subColorsZn = ['#2563eb','#f59e0b','#10b981','#ef4444','#8b5cf6'];
    var subVolsZn = actualSubsegsZn.map(function(s){return tsVols(getZoneIndustryVols(zone,s));});
    var mixTotalsZn = subVolsZn[0].map(function(_,t){return subVolsZn.reduce(function(sum,sv){return sum+sv[t];},0);});
    var subPctsZn = subVolsZn.map(function(sv){return sv.map(function(v,t){return mixTotalsZn[t]>0?v/mixTotalsZn[t]*100:0;});});
    var datesZn = tsDates(), tsLenZn = datesZn.length;

    document.getElementById('title-zn-subseg-mix').textContent = segNameZn + ' Subsegment Mix (%) - ' + zone;
    var mixTracesZn = actualSubsegsZn.map(function(s,i){return {
      x:datesZn, y:subPctsZn[i], customdata:subVolsZn[i], type:'bar', name:s,
      marker:{color:subColorsZn[i%subColorsZn.length]},
      hovertemplate:s+': %{y:.1f}%<br>Volume: %{customdata:,.0f}<extra></extra>'
    };});
    Plotly.newPlot('chart-zn-subseg-mix',mixTracesZn,{
      ...PLOTLY_LAYOUT, barmode:'stack',
      margin:{l:60,r:25,t:10,b:75},
      yaxis:{...PLOTLY_LAYOUT.yaxis, title:{text:'% Mix',standoff:10}, range:[0,100.5]},
      legend:{orientation:'h',y:-0.22,x:0.5,xanchor:'center',font:{size:10}}
    },PLOTLY_CONFIG);
    addChartCopyBtn('chart-zn-subseg-mix');

    document.getElementById('title-zn-subseg-vol').textContent = segNameZn + ' Subsegment Volumes - ' + zone;
    var volHeadZn = '<tr><th>Subsegment</th>' + datesZn.map(function(d){return '<th class="align-right">'+d+'</th>';}).join('') + '</tr>';
    var volBodyZn = '';
    actualSubsegsZn.forEach(function(s,i){
      volBodyZn += '<tr><td><b>'+s+'</b></td>';
      for(var t=0;t<tsLenZn;t++) volBodyZn += '<td class="align-right">'+fmt(subVolsZn[i][t])+'</td>';
      volBodyZn += '</tr>';
    });
    volBodyZn += '<tr class="total-row"><td><b>Total '+segNameZn+'</b></td>';
    for(var t=0;t<tsLenZn;t++) volBodyZn += '<td class="align-right"><b>'+fmt(mixTotalsZn[t])+'</b></td>';
    volBodyZn += '</tr>';
    actualSubsegsZn.forEach(function(s,i){
      volBodyZn += '<tr style="color:#6b7280;font-size:12px"><td><i>'+s+' %</i></td>';
      for(var t=0;t<tsLenZn;t++) volBodyZn += '<td class="align-right"><i>'+subPctsZn[i][t].toFixed(1)+'%</i></td>';
      volBodyZn += '</tr>';
    });
    document.querySelector('#table-zn-subseg-vol thead').innerHTML = volHeadZn;
    document.querySelector('#table-zn-subseg-vol tbody').innerHTML = volBodyZn;

    document.getElementById('title-zn-subseg-yoy').textContent = segNameZn + ' Subsegment YoY Growth (%) - ' + zone;
    var yoyHeadZn = '<tr><th>Subsegment</th>' + datesZn.map(function(d){return '<th class="align-right">'+d+'</th>';}).join('') + '</tr>';
    var yoyBodyZn = '';
    actualSubsegsZn.forEach(function(s){
      var yoySZ = yoyGrowthSeries(getZoneIndustryVols(zone,s));
      yoyBodyZn += '<tr><td><b>'+s+'</b></td>';
      for(var t=0;t<tsLenZn;t++){
        var gZ=yoySZ[t];
        if(gZ===null||gZ===undefined) yoyBodyZn+='<td class="align-right">-</td>';
        else yoyBodyZn+='<td class="align-right '+(gZ>=0?'positive':'negative')+'">'+(gZ>=0?'+':'')+gZ.toFixed(1)+'%</td>';
      }
      yoyBodyZn += '</tr>';
    });
    var overallYoYZn = yoyGrowthSeries(znIndQ);
    yoyBodyZn += '<tr class="total-row"><td><b>Overall '+segNameZn+'</b></td>';
    for(var t=0;t<tsLenZn;t++){
      var g2Z=overallYoYZn[t];
      if(g2Z===null||g2Z===undefined) yoyBodyZn+='<td class="align-right">-</td>';
      else yoyBodyZn+='<td class="align-right '+(g2Z>=0?'positive':'negative')+'"><b>'+(g2Z>=0?'+':'')+g2Z.toFixed(1)+'%</b></td>';
    }
    yoyBodyZn += '</tr>';
    document.querySelector('#table-zn-subseg-yoy thead').innerHTML = yoyHeadZn;
    document.querySelector('#table-zn-subseg-yoy tbody').innerHTML = yoyBodyZn;
  } else {
    mixElZn.style.display = 'none';
  }

  // State breakdown within zone (horizontal bar)
  const statesInZone = zoneStateMap[zone] || [];
  const stateVols = {};
  for (const st of statesInZone) {
    const sv = getStateIndustryVols(st, sub);
    const vol = pQs.reduce((s,qi)=>s+(sv[qi]||0),0);
    if (vol > 0) stateVols[st] = vol;
  }
  const sortedStates = Object.entries(stateVols).sort((a,b)=>b[1]-a[1]);
  plotHBar('chart-zn-states',
    sortedStates.map(e=>e[0]).reverse(),
    sortedStates.map(e=>e[1]).reverse(),
    sortedStates.map(()=>'#2563eb').reverse()
  );

  // State contribution trend (top states within zone)
  const topStates = sortedStates.slice(0,10).map(e=>e[0]);
  const znTotalTS = tsV; // zone total volumes time-series
  const stContribTraces = [];
  topStates.forEach((st,si) => {
    const stQ = getStateIndustryVols(st, sub);
    const stTS = tsVols(stQ);
    const contribTS = stTS.map((v,t) => znTotalTS[t]>0 ? v/znTotalTS[t]*100 : 0);
    stContribTraces.push({
      x:tsDates(), y:contribTS, type:'scatter', mode:'lines+markers',
      name:st, line:{color:PALETTE[si%PALETTE.length],width:2}, marker:{size:3},
      hovertemplate:st+': %{y:.1f}%<extra></extra>'
    });
  });
  if (sortedStates.length > 10) {
    const othersStates = sortedStates.slice(10).map(e=>e[0]);
    const othersContrib = tsDates().map((_,t) => {
      const topSum = topStates.reduce((s,st) => s + tsVols(getStateIndustryVols(st,sub))[t], 0);
      return znTotalTS[t]>0 ? (znTotalTS[t]-topSum)/znTotalTS[t]*100 : 0;
    });
    stContribTraces.push({
      x:tsDates(), y:othersContrib, type:'scatter', mode:'lines',
      name:'Others', line:{color:'#d1d5db',width:1.5,dash:'dot'},
      hovertemplate:'Others: %{y:.1f}%<extra></extra>'
    });
  }
  plotLines('chart-zn-state-contrib', stContribTraces, 'Contribution (%)', true);

  // State details table
  const yoyQs = yoyPeriodQIdxs();
  const stateDetails = sortedStates.map(([st, vol]) => {
    const stIndQ = getStateIndustryVols(st, sub);
    const sVol = pQs.reduce((s,qi)=>s+(stIndQ[qi]||0),0);
    const sYoyVol = yoyQs.reduce((s,qi)=>s+(stIndQ[qi]||0),0);
    const sGrowth = sYoyVol>0?((sVol/sYoyVol-1)*100):0;
    const sShare = curVol>0?sVol/curVol*100:0;
    const znIndYoy = yoyQs.reduce((s,qi)=>s+(znIndQ[qi]||0),0);
    const sShareYoy = znIndYoy>0?sYoyVol/znIndYoy*100:0;
    return {state:st,vol:sVol,yoyVol:sYoyVol,growth:sGrowth,share:sShare,shareChg:sShare-sShareYoy,contrib:curVol>0?sVol/curVol*100:0};
  }).filter(d=>d.vol>0);
  const stTotVol=stateDetails.reduce((s,d)=>s+d.vol,0);
  const stTotYoy=stateDetails.reduce((s,d)=>s+d.yoyVol,0);
  const stTotG=stTotYoy>0?((stTotVol/stTotYoy-1)*100):0;

  const stThead = document.querySelector('#table-zn-states thead');
  stThead.innerHTML = '<tr><th>#</th><th>State</th><th class="align-right">Volume</th><th class="align-right">YoY Vol</th><th class="align-right">YoY Growth</th><th class="align-right">Share of Zone</th><th class="align-right">Share Chg</th></tr>';
  const stTbody = document.querySelector('#table-zn-states tbody');
  stTbody.innerHTML = stateDetails.map((d,i) => `
    <tr class="clickable" onclick="drillToState('${esc(d.state)}','${esc(zone)}')">
      <td>${i+1}</td><td><b>${d.state}</b></td>
      <td class="align-right">${fmt(d.vol)}</td><td class="align-right">${fmt(d.yoyVol)}</td>
      <td class="align-right ${d.growth>=0?'positive':'negative'}">${d.growth>=0?'+':''}${d.growth.toFixed(1)}%</td>
      <td class="align-right">${d.share.toFixed(1)}%</td>
      <td class="align-right"><span class="badge ${d.shareChg>=0?'badge-green':'badge-red'}">${fmtPP(d.shareChg)}</span></td>
    </tr>`).join('') + `
    <tr class="total-row"><td></td><td><b>TOTAL</b></td>
      <td class="align-right"><b>${fmt(stTotVol)}</b></td><td class="align-right"><b>${fmt(stTotYoy)}</b></td>
      <td class="align-right ${stTotG>=0?'positive':'negative'}"><b>${stTotG>=0?'+':''}${stTotG.toFixed(1)}%</b></td>
      <td class="align-right"><b>100.0%</b></td><td class="align-right"></td>
    </tr>`;

  // Company rankings table within zone
  const comps = topCompanies(50, sub, null, zone);
  const compTableData = comps.map(c => {
    const cv = getZoneCompanyVols(zone, c, sub);
    const v = pQs.reduce((s,qi)=>s+(cv[qi]||0),0);
    const yv = yoyQs.reduce((s,qi)=>s+(cv[qi]||0),0);
    const g = yv>0?((v/yv-1)*100):0;
    const s = curVol>0?v/curVol*100:0;
    const znIndYoy2 = yoyQs.reduce((sum,qi)=>sum+(znIndQ[qi]||0),0);
    const sy = znIndYoy2>0?yv/znIndYoy2*100:0;
    return {company:c,vol:v,yoyVol:yv,growth:g,share:s,shareChg:s-sy};
  }).filter(d=>d.vol>0);
  compTableData.sort((a,b)=>b.vol-a.vol);
  const cTotVol=compTableData.reduce((s,d)=>s+d.vol,0);
  const cTotYoy=compTableData.reduce((s,d)=>s+d.yoyVol,0);
  const cTotG=cTotYoy>0?((cTotVol/cTotYoy-1)*100):0;

  const cThead = document.querySelector('#table-zn-companies thead');
  cThead.innerHTML = '<tr><th>#</th><th>Company</th><th class="align-right">Volume</th><th class="align-right">YoY Vol</th><th class="align-right">YoY Growth</th><th class="align-right">Mkt Share</th><th class="align-right">Share Chg</th></tr>';
  const cTbody = document.querySelector('#table-zn-companies tbody');
  cTbody.innerHTML = compTableData.map((d,i) => `
    <tr class="clickable" onclick="drillToCompany('${esc(d.company)}')">
      <td>${i+1}</td><td><b>${d.company}</b></td>
      <td class="align-right">${fmt(d.vol)}</td><td class="align-right">${fmt(d.yoyVol)}</td>
      <td class="align-right ${d.growth>=0?'positive':'negative'}">${d.growth>=0?'+':''}${d.growth.toFixed(1)}%</td>
      <td class="align-right">${d.share.toFixed(1)}%</td>
      <td class="align-right"><span class="badge ${d.shareChg>=0?'badge-green':'badge-red'}">${fmtPP(d.shareChg)}</span></td>
    </tr>`).join('') + `
    <tr class="total-row"><td></td><td><b>TOTAL</b></td>
      <td class="align-right"><b>${fmt(cTotVol)}</b></td><td class="align-right"><b>${fmt(cTotYoy)}</b></td>
      <td class="align-right ${cTotG>=0?'positive':'negative'}"><b>${cTotG>=0?'+':''}${cTotG.toFixed(1)}%</b></td>
      <td class="align-right"><b>100.0%</b></td><td class="align-right"></td>
    </tr>`;
}

// ============================================
// UI HELPERS
// ============================================
function esc(s) { return s.replace(/'/g,"\\\\'"); }

function switchSegment(seg) {
  currentSegment = seg;
  currentSubseg = 'All';
  document.querySelectorAll('.seg-tab').forEach(t => t.classList.toggle('active',t.dataset.seg===seg));
  buildIndexes();
  updateDropdowns();
  renderSubsegChips();
  populatePeriodSelector(currentTab);
  renderCurrentTab();
}

function switchTab(tab) {
  currentTab = tab;
  currentSubseg = 'All';
  document.querySelectorAll('.nav-tab').forEach(t => t.classList.toggle('active',t.dataset.tab===tab));
  document.querySelectorAll('.panel').forEach(p => p.classList.remove('active'));
  document.getElementById('panel-'+tab).classList.add('active');
  renderSubsegChips();
  syncViewChips();
  populatePeriodSelector(tab);
  renderCurrentTab();
}

function renderSubsegChips() {
  const el = document.getElementById('subsegChips-'+currentTab);
  if (!el) return;
  el.innerHTML = '<div class="subseg-chip active" data-sub="All">All</div>' +
    segSubsegs.map(s=>`<div class="subseg-chip" data-sub="${s}">${s}</div>`).join('');
  el.querySelectorAll('.subseg-chip').forEach(chip => {
    chip.onclick = () => {
      currentSubseg = chip.dataset.sub;
      el.querySelectorAll('.subseg-chip').forEach(c=>c.classList.toggle('active',c.dataset.sub===currentSubseg));
      renderCurrentTab();
    };
  });
}

function syncViewChips() {
  const el = document.getElementById('viewChips-'+currentTab);
  if (!el) return;
  el.querySelectorAll('.view-chip').forEach(c => c.classList.toggle('active',c.dataset.view===viewModes[currentTab]));
}

function updateDropdowns() {
  const selCo = document.getElementById('sel-company');
  const topC = topCompanies(50,'All',null);
  selCo.innerHTML = topC.map(c=>`<option value="${c}" ${c===currentCompany?'selected':''}>${c}</option>`).join('');
  if (!currentCompany||!topC.includes(currentCompany)) currentCompany=topC[0]||'';
  selCo.value = currentCompany;
  const selZone = document.getElementById('sel-zone');
  selZone.innerHTML = '<option value="All">All Zones</option>'+segZones.map(z=>`<option value="${z}">${z}</option>`).join('');
  selZone.value = currentZone;
  updateStateDropdown();
  updateZoneTabDropdown();
}

function updateZoneTabDropdown() {
  const sel = document.getElementById('sel-zone-tab');
  sel.innerHTML = segZones.map(z=>`<option value="${z}" ${z===currentZoneTab?'selected':''}>${z}</option>`).join('');
  if (!currentZoneTab || !segZones.includes(currentZoneTab)) currentZoneTab = segZones[0] || '';
  sel.value = currentZoneTab;
}

function updateStateDropdown() {
  const selState = document.getElementById('sel-state');
  const states = currentZone==='All'?segStates:(zoneStateMap[currentZone]||[]);
  selState.innerHTML = states.map(s=>`<option value="${s}" ${s===currentState?'selected':''}>${s}</option>`).join('');
  if (!currentState||!states.includes(currentState)) currentState=states[0]||'';
  selState.value = currentState;
}

function renderCurrentTab() {
  if (currentTab==='overview') renderOverview();
  else if (currentTab==='company') renderCompanyView();
  else if (currentTab==='state') renderStateView();
  else if (currentTab==='zone') renderZoneView();
  else if (currentTab==='chat') renderChatTab();
  else if (currentTab==='data') renderDataTab();
}

function drillToCompany(company) {
  currentCompany = company;
  document.getElementById('sel-company').value = company;
  switchTab('company');
}
function drillToState(state, zone) {
  if (zone && zone!=='All'){currentZone=zone;document.getElementById('sel-zone').value=zone;updateStateDropdown();}
  currentState = state;
  document.getElementById('sel-state').value = state;
  switchTab('state');
}
function drillToZone(zone) {
  currentZoneTab = zone;
  document.getElementById('sel-zone-tab').value = zone;
  switchTab('zone');
}

// ============================================
// DATA MANAGEMENT
// ============================================
function renderDataTab() {
  const infoGrid = document.getElementById('data-info-grid');
  const resetBtn = document.getElementById('btn-reset-data');
  const statusEl = document.getElementById('data-status');
  const segments = [...new Set(ROWS.map(r => r[0]))];
  const companies = [...new Set(ROWS.map(r => r[4]))];
  const states = [...new Set(ROWS.map(r => r[3]))];
  const meta = JSON.parse(localStorage.getItem("janchor_auto_meta") || "null");
  infoGrid.innerHTML =
    '<div class="data-info-item"><div class="di-label">Quarters</div><div class="di-value">'+NQ+'</div></div>'+
    '<div class="data-info-item"><div class="di-label">Period Range</div><div class="di-value">'+QLABELS[0]+' \u2013 '+QLABELS[NQ-1]+'</div></div>'+
    '<div class="data-info-item"><div class="di-label">Segments</div><div class="di-value">'+segments.length+'</div></div>'+
    '<div class="data-info-item"><div class="di-label">Companies</div><div class="di-value">'+companies.length+'</div></div>'+
    '<div class="data-info-item"><div class="di-label">States</div><div class="di-value">'+states.length+'</div></div>'+
    '<div class="data-info-item"><div class="di-label">Total Rows</div><div class="di-value">'+ROWS.length.toLocaleString()+'</div></div>'+
    '<div class="data-info-item"><div class="di-label">Data Source</div><div class="di-value">'+(DATA_IS_CUSTOM ? 'Uploaded File' : 'Embedded (Default)')+'</div></div>'+
    (meta ? '<div class="data-info-item"><div class="di-label">Uploaded</div><div class="di-value">'+new Date(meta.uploadedAt).toLocaleDateString()+'</div></div>' : '')+
    (meta && meta.fileName ? '<div class="data-info-item"><div class="di-label">File Name</div><div class="di-value">'+meta.fileName+'</div></div>' : '');
  resetBtn.style.display = DATA_IS_CUSTOM ? 'inline-flex' : 'none';
  if (DATA_IS_CUSTOM) {
    showStatus('info', 'Using uploaded data. Click "Reset to Embedded Data" to revert to the original dataset.');
  } else {
    statusEl.className = 'status-msg';
  }
}

function showStatus(type, msg) {
  var el = document.getElementById('data-status');
  el.className = 'status-msg show ' + type;
  el.textContent = msg;
}

/* OEM name cleanup: map full legal names (Kotak format) to short display names */
function cleanOemName(name) {
  var M = {
    'Maruti Suzuki India Ltd':'Maruti Suzuki','Hyundai Motor India Ltd':'Hyundai',
    'Tata Motors Ltd':'Tata Motors','Mahindra & Mahindra Ltd':'M&M',
    'Toyota Kirloskar Motor Pvt Ltd':'Toyota','Kia Motors':'Kia',
    'Honda Cars India Ltd':'Honda','SkodaAuto India Pvt Ltd':'Skoda',
    'Volkswagen India Pvt Ltd':'Volkswagen','Renault India Pvt Ltd':'Renault',
    'Nissan Motor India Pvt Ltd':'Nissan','Ford India Pvt Ltd':'Ford',
    'Fiat India Automobiles Pvt.Ltd':'Fiat India',
    'General Motors India Pvt Ltd':'General Motors',
    'Hindustan Motor Finance Corporation Ltd':'Hindustan Motor',
    'Force Motors Ltd':'Force Motors','Isuzu Motors India Pvt Ltd':'Isuzu',
    'PCA Motors Pvt Ltd':'PCA Motors','MG Motor':'MG Motor',
    'Hero MotoCorp Ltd':'Hero',
    'Honda Motorcycle & Scooter India (Pvt) Ltd':'Honda',
    'TVS Motor Company Ltd':'TVS Motor','Bajaj Auto Ltd':'Bajaj',
    'Royal Enfield (A Unit of Eicher Motors Ltd)':'Royal Enfield',
    'Suzuki Motorcycle India Pvt Ltd':'Suzuki',
    'India Yamaha Motor Pvt Ltd':'Yamaha',
    'Ather Energy Pvt. Ltd':'Ather','Okinawa Autotech Pvt. Ltd':'Okinawa',
    'India Kawasaki Motors Pvt Ltd':'Kawasaki',
    'H-D Motor Company India Pvt Ltd':'Harley Davidson',
    'Mahindra Two Wheelers Ltd':'M&M',
    'Piaggio Vehicles Pvt Ltd':'Piaggio',
    'UM Lohia Two Wheelers Pvt Ltd':'UM Lohia',
    'Ashok Leyland Ltd':'Ashok Leyland','VECV- Eicher':'VECV',
    'Volvo group':'Volvo Group','SML Isuzu Ltd':'SML Isuzu',
    'TI Clean Mobility Pvt Ltd':'TI Clean Mobility',
    'Pinnacle Mobility Solutions Pvt Ltd':'Pinnacle Mobility',
    'Atul Auto Ltd':'Atul Auto'
  };
  if (M[name]) return M[name];
  return name.replace(/\\s*\\(Pvt\\)\\s*/gi,' ').replace(/\\s*(Pvt\\.?|Private)\\s*(Ltd\\.?|Limited)\\s*$/i,'').replace(/\\s*(Ltd\\.?|Limited)\\s*$/i,'').replace(/\\s+India\\s*$/i,'').replace(/\\s+/g,' ').trim()||name;
}

/* Auto-detect Excel format and dispatch to appropriate parser */
function parseExcel(arrayBuffer, fileName) {
  try {
    showStatus('info', 'Parsing Excel file...');
    var wb = XLSX.read(arrayBuffer, {type: 'array'});
    /* Detect old format: has sheets like "PVs - Raw data" */
    var OLD_SHEETS = ['PVs - Raw data','2Ws - Raw data','3Ws - Raw data','M&HCVs - Raw data','LCVs - Raw data'];
    var isOldFormat = OLD_SHEETS.some(function(s){ return wb.SheetNames.indexOf(s) >= 0; });
    if (isOldFormat) return parseOldFormat(wb);
    /* Detect new Kotak format: has sheets like "Cars","UVs","Motorcycle" etc. */
    var NEW_SHEETS = ['Cars','UVs','Motorcycle','Scooters','MHCVs','LCVs','3W','PVs','2W'];
    var isNewFormat = NEW_SHEETS.some(function(s){ return wb.SheetNames.indexOf(s) >= 0; });
    if (isNewFormat) return parseNewFormat(wb);
    showStatus('error', 'Unrecognized Excel format. Expected sheets like "PVs - Raw data" (old format) or "Cars","UVs" etc. (Kotak format). Found: ' + wb.SheetNames.join(', '));
    return null;
  } catch(e) {
    showStatus('error', 'Error parsing Excel: ' + e.message);
    return null;
  }
}

/* Parser for old format with explicit Zone/State/Manufacturer columns */
function parseOldFormat(wb) {
  var SHEET_MAP = {
    'PVs - Raw data': 'PV', '2Ws - Raw data': '2W', '3Ws - Raw data': '3W',
    'M&HCVs - Raw data': 'MHCV', 'LCVs - Raw data': 'LCV'
  };
  var allQuarters = null, allRows = [], sheetsFound = 0;
  var sheetNames = Object.keys(SHEET_MAP);
  for (var si = 0; si < sheetNames.length; si++) {
    var sheetName = sheetNames[si];
    var segment = SHEET_MAP[sheetName];
    if (wb.SheetNames.indexOf(sheetName) < 0) {
      showStatus('error', 'Missing sheet: "' + sheetName + '". Please check your Excel file has all 5 raw data sheets.');
      return null;
    }
    sheetsFound++;
    var ws = wb.Sheets[sheetName];
    var jsonData = XLSX.utils.sheet_to_json(ws, {header:1, defval:0});
    if (jsonData.length < 3) { showStatus('error', 'Sheet "' + sheetName + '" has insufficient data.'); return null; }
    var headerRowIdx = -1;
    for (var r = 0; r < Math.min(5, jsonData.length); r++) {
      var row = jsonData[r];
      for (var c = 0; c < row.length; c++) {
        var val = String(row[c] || '');
        if (/^Q\\dFY\\d{2}$/.test(val)) { headerRowIdx = r; break; }
      }
      if (headerRowIdx >= 0) break;
    }
    if (headerRowIdx < 0) { showStatus('error', 'Cannot find quarter headers in sheet "' + sheetName + '".'); return null; }
    var headers = jsonData[headerRowIdx];
    var colZone=-1, colState=-1, colMfr=-1, colSubseg=-1, qStartCol=-1, quarters=[];
    for (var c2 = 0; c2 < headers.length; c2++) {
      var h = String(headers[c2] || '').trim(), hLow = h.toLowerCase();
      if (hLow === 'zone') colZone = c2;
      else if (hLow === 'state') colState = c2;
      else if (hLow === 'manufacturer' || hLow === 'oem') colMfr = c2;
      else if (hLow === 'sub-segment' || hLow === 'subsegment' || hLow === 'sub_segment' || hLow === 'sub segment') colSubseg = c2;
      else if (/^Q\\dFY\\d{2}$/.test(h)) { if (qStartCol < 0) qStartCol = c2; quarters.push(h); }
    }
    if (colZone < 0 || colState < 0 || colMfr < 0 || qStartCol < 0) {
      showStatus('error', 'Cannot find required columns (Zone, State, Manufacturer, Quarters) in sheet "' + sheetName + '".'); return null;
    }
    if (allQuarters === null) allQuarters = quarters;
    else if (quarters.length > allQuarters.length) allQuarters = quarters;
    for (var r2 = headerRowIdx + 1; r2 < jsonData.length; r2++) {
      var drow = jsonData[r2];
      var zone = String(drow[colZone] || '').trim();
      var state = String(drow[colState] || '').trim();
      var mfr = String(drow[colMfr] || '').trim();
      var subseg = colSubseg >= 0 ? String(drow[colSubseg] || '').trim() : 'All';
      if (!zone || !state || !mfr) continue;
      if (zone.toLowerCase() === 'zone' || state.toLowerCase() === 'state') continue;
      var volumes = [], hasNonZero = false;
      for (var q = 0; q < quarters.length; q++) {
        var v = drow[qStartCol + q];
        var num = typeof v === 'number' ? v : (parseFloat(v) || 0);
        volumes.push(num); if (num > 0) hasNonZero = true;
      }
      if (hasNonZero) allRows.push([segment, subseg, zone, state, mfr].concat(volumes));
    }
  }
  if (allRows.length === 0) { showStatus('error', 'No valid data rows found.'); return null; }
  var result = {
    quarters: allQuarters,
    columns: ['segment','subsegment','zone','state','manufacturer'].concat(allQuarters.map(function(q){return 'vol_'+q.toLowerCase();})),
    rows: allRows
  };
  showStatus('success', 'Parsed ' + allRows.length.toLocaleString() + ' rows across ' + sheetsFound + ' segments, ' + allQuarters.length + ' quarters (' + allQuarters[0] + ' to ' + allQuarters[allQuarters.length-1] + '). Saving...');
  return result;
}

/* Parser for new Kotak hierarchical format (Zone > State > OEM structure) */
function parseNewFormat(wb) {
  var ZONE_NAMES = ['North Zone','East Zone','West Zone','South Zone'];
  var ZONE_SHORT = {'North Zone':'North','East Zone':'East','West Zone':'West','South Zone':'South'};
  /* Sheet mapping: subsegment sheets take priority over parent aggregate sheets */
  var sheetDefs = [];
  var hasCars = wb.SheetNames.indexOf('Cars') >= 0;
  var hasUVs = wb.SheetNames.indexOf('UVs') >= 0;
  var hasMotorcycle = wb.SheetNames.indexOf('Motorcycle') >= 0;
  var hasScooters = wb.SheetNames.indexOf('Scooters') >= 0;
  if (hasCars) sheetDefs.push({sheet:'Cars',seg:'PV',sub:'Cars'});
  if (hasUVs) sheetDefs.push({sheet:'UVs',seg:'PV',sub:'UVs'});
  if (!hasCars && !hasUVs && wb.SheetNames.indexOf('PVs') >= 0) sheetDefs.push({sheet:'PVs',seg:'PV',sub:'All'});
  if (hasMotorcycle) sheetDefs.push({sheet:'Motorcycle',seg:'2W',sub:'Motorcycle'});
  if (hasScooters) sheetDefs.push({sheet:'Scooters',seg:'2W',sub:'Scooters'});
  if (!hasMotorcycle && !hasScooters && wb.SheetNames.indexOf('2W') >= 0) sheetDefs.push({sheet:'2W',seg:'2W',sub:'All'});
  if (wb.SheetNames.indexOf('MHCVs') >= 0) sheetDefs.push({sheet:'MHCVs',seg:'MHCV',sub:'All'});
  if (wb.SheetNames.indexOf('LCVs') >= 0) sheetDefs.push({sheet:'LCVs',seg:'LCV',sub:'All'});
  if (wb.SheetNames.indexOf('3W') >= 0) sheetDefs.push({sheet:'3W',seg:'3W',sub:'All'});

  var allQuarters = null, allRows = [], sheetsFound = 0;
  for (var si = 0; si < sheetDefs.length; si++) {
    var def = sheetDefs[si];
    var ws = wb.Sheets[def.sheet];
    if (!ws) continue;
    sheetsFound++;
    var jsonData = XLSX.utils.sheet_to_json(ws, {header:1, defval:0});
    if (jsonData.length < 5) continue;

    /* Find header row with quarter pattern like 1QFY16 */
    var headerRowIdx = -1;
    for (var r = 0; r < Math.min(10, jsonData.length); r++) {
      var row = jsonData[r];
      for (var c = 0; c < (row ? row.length : 0); c++) {
        var val = String(row[c] || '');
        if (/^\\dQFY\\d{2}$/.test(val)) { headerRowIdx = r; break; }
      }
      if (headerRowIdx >= 0) break;
    }
    if (headerRowIdx < 0) continue;

    var headers = jsonData[headerRowIdx];
    /* Find name column and quarter columns */
    var nameCol = -1, qStartCol = -1, quarters = [];
    for (var c2 = 0; c2 < headers.length; c2++) {
      var h = String(headers[c2] || '').trim();
      if (/^\\dQFY\\d{2}$/.test(h)) {
        if (qStartCol < 0) qStartCol = c2;
        quarters.push('Q' + h.charAt(0) + h.substring(2)); /* 1QFY16 -> Q1FY16 */
      }
    }
    /* Detect name column by finding first zone name in rows after header */
    for (var r3 = headerRowIdx + 1; r3 < Math.min(headerRowIdx + 15, jsonData.length); r3++) {
      var testRow = jsonData[r3];
      if (!testRow) continue;
      for (var c3 = 0; c3 < Math.min(5, testRow.length); c3++) {
        var tv = String(testRow[c3] || '').trim();
        if (ZONE_NAMES.indexOf(tv) >= 0) { nameCol = c3; break; }
      }
      if (nameCol >= 0) break;
    }
    if (nameCol < 0 || qStartCol < 0 || quarters.length === 0) continue;

    if (allQuarters === null) allQuarters = quarters;
    else if (quarters.length > allQuarters.length) allQuarters = quarters;

    /* Detect OEM list using first-repeat algorithm:
       After first zone row, collect names. When a name repeats, everything
       before the last name = OEMs, the last name = first state. */
    var oemList = [], oemSet = {}, oemDetected = false;
    var firstZoneRow = -1;
    for (var r4 = headerRowIdx + 1; r4 < jsonData.length; r4++) {
      var n4 = String(jsonData[r4][nameCol] || '').trim();
      if (ZONE_NAMES.indexOf(n4) >= 0) { firstZoneRow = r4; break; }
    }
    if (firstZoneRow < 0) continue;

    var seenNames = {};
    for (var r5 = firstZoneRow + 1; r5 < jsonData.length; r5++) {
      var n5 = String(jsonData[r5][nameCol] || '').trim();
      if (!n5 || n5 === '0') continue;
      if (seenNames[n5]) { oemDetected = true; break; }
      seenNames[n5] = true;
      oemList.push(n5);
    }
    if (!oemDetected || oemList.length < 2) continue;

    /* Last name before repeat = first state; rest are OEMs */
    oemList.pop();
    for (var oi = 0; oi < oemList.length; oi++) oemSet[oemList[oi]] = true;

    /* Parse all data rows with state machine */
    var currentZone = '', currentState = '', inZoneBlock = false;
    for (var r6 = firstZoneRow; r6 < jsonData.length; r6++) {
      var drow = jsonData[r6];
      var name6 = String(drow[nameCol] || '').trim();
      if (!name6 || name6 === '0') continue;

      /* Stop at Total / Grand Total / All India */
      var nameLow = name6.toLowerCase();
      if (nameLow === 'total' || nameLow === 'grand total' || nameLow === 'all india' || nameLow === 'india') break;

      /* Zone row */
      if (ZONE_NAMES.indexOf(name6) >= 0) {
        currentZone = ZONE_SHORT[name6] || name6;
        currentState = '';
        inZoneBlock = true;
        continue;
      }

      /* OEM row */
      if (oemSet[name6]) {
        if (inZoneBlock) continue; /* skip zone-level OEM aggregates */
        if (!currentZone || !currentState) continue;
        var volumes = [], hasNonZero = false;
        for (var q = 0; q < quarters.length; q++) {
          var v = drow[qStartCol + q];
          var num = typeof v === 'number' ? v : (parseFloat(v) || 0);
          volumes.push(num); if (num > 0) hasNonZero = true;
        }
        if (hasNonZero) {
          allRows.push([def.seg, def.sub, currentZone, currentState, cleanOemName(name6)].concat(volumes));
        }
        continue;
      }

      /* State name row */
      currentState = name6;
      inZoneBlock = false;
    }
  }
  if (allRows.length === 0) { showStatus('error', 'No valid data rows found in the Kotak file.'); return null; }
  var result = {
    quarters: allQuarters,
    columns: ['segment','subsegment','zone','state','manufacturer'].concat(allQuarters.map(function(q){return 'vol_'+q.toLowerCase();})),
    rows: allRows
  };
  showStatus('success', 'Parsed ' + allRows.length.toLocaleString() + ' rows across ' + sheetsFound + ' segments (' + sheetDefs.map(function(d){return d.sheet;}).join(', ') + '), ' + allQuarters.length + ' quarters (' + allQuarters[0] + ' to ' + allQuarters[allQuarters.length-1] + '). Saving...');
  return result;
}

function handleFileUpload(file) {
  if (!file) return;
  var ext = file.name.split('.').pop().toLowerCase();
  if (ext !== 'xlsx' && ext !== 'xls') {
    showStatus('error', 'Please upload an Excel file (.xlsx or .xls)');
    return;
  }
  showStatus('info', 'Reading file: ' + file.name + '...');
  var reader = new FileReader();
  reader.onload = function(e) {
    var data = parseExcel(e.target.result, file.name);
    if (data) {
      try {
        var jsonStr = JSON.stringify(data);
        localStorage.setItem("janchor_auto_data", jsonStr);
        localStorage.setItem("janchor_auto_meta", JSON.stringify({
          fileName: file.name,
          uploadedAt: new Date().toISOString(),
          rows: data.rows.length,
          quarters: data.quarters.length
        }));
        showStatus('success', 'Data saved! Reloading dashboard with new data...');
        setTimeout(function(){ location.reload(); }, 1500);
      } catch(err) {
        if (err.name === 'QuotaExceededError') {
          showStatus('error', 'Data too large for browser storage. The file exceeds the 5MB browser storage limit.');
        } else {
          showStatus('error', 'Error saving data: ' + err.message);
        }
      }
    }
  };
  reader.onerror = function() {
    showStatus('error', 'Error reading file. Please try again.');
  };
  reader.readAsArrayBuffer(file);
}

function resetData() {
  localStorage.removeItem("janchor_auto_data");
  localStorage.removeItem("janchor_auto_meta");
  showStatus('success', 'Data reset! Reloading with embedded data...');
  setTimeout(function(){ location.reload(); }, 1000);
}

// ============================================
// TABLE SORTING
// ============================================
document.addEventListener('click', function(e) {
  if (e.target.tagName==='TH') {
    const table=e.target.closest('table'), tbody=table.querySelector('tbody');
    const idx=Array.from(e.target.parentNode.children).indexOf(e.target);
    const rows=Array.from(tbody.querySelectorAll('tr:not(.total-row)'));
    const isAsc=e.target.dataset.sort==='asc';
    rows.sort((a,b) => {
      let va=a.children[idx]?.textContent.trim()||'';
      let vb=b.children[idx]?.textContent.trim()||'';
      const na=parseFloat(va.replace(/[^\\d.\\-]/g,''));
      const nb=parseFloat(vb.replace(/[^\\d.\\-]/g,''));
      if (!isNaN(na)&&!isNaN(nb)) return isAsc?na-nb:nb-na;
      return isAsc?va.localeCompare(vb):vb.localeCompare(va);
    });
    e.target.dataset.sort = isAsc?'desc':'asc';
    const totalRow = tbody.querySelector('.total-row');
    rows.forEach(r=>tbody.appendChild(r));
    if (totalRow) tbody.appendChild(totalRow);
  }
});

// ============================================
// EVENT LISTENERS
// ============================================
document.querySelectorAll('.seg-tab').forEach(t=>t.onclick=()=>switchSegment(t.dataset.seg));
document.querySelectorAll('.nav-tab').forEach(t=>t.onclick=()=>switchTab(t.dataset.tab));
document.getElementById('sel-company').onchange=function(){currentCompany=this.value;renderCurrentTab();};
document.getElementById('sel-zone').onchange=function(){currentZone=this.value;updateStateDropdown();renderCurrentTab();};
document.getElementById('sel-state').onchange=function(){currentState=this.value;renderCurrentTab();};
document.getElementById('sel-zone-tab').onchange=function(){currentZoneTab=this.value;renderCurrentTab();};

// View mode toggles
document.querySelectorAll('[id^="viewChips-"]').forEach(container => {
  container.querySelectorAll('.view-chip').forEach(chip => {
    chip.onclick = () => {
      const tab = container.id.replace('viewChips-','');
      viewModes[tab] = chip.dataset.view;
      // Reset period to latest
      selectedPeriods[tab] = chip.dataset.view==='annual' ? NFY-1 : NQ-1;
      container.querySelectorAll('.view-chip').forEach(c=>c.classList.toggle('active',c.dataset.view===chip.dataset.view));
      populatePeriodSelector(tab);
      if (tab===currentTab) renderCurrentTab();
    };
  });
});

// Period selectors
['overview','company','state','zone'].forEach(tab => {
  const sel = document.getElementById('sel-period-'+tab);
  if (sel) sel.onchange = function() {
    selectedPeriods[tab] = parseInt(this.value);
    if (tab===currentTab) renderCurrentTab();
  };
});

// Data management event listeners
(function() {
  var uploadZone = document.getElementById('upload-zone');
  var fileInput = document.getElementById('file-input');
  var resetBtn = document.getElementById('btn-reset-data');
  if (uploadZone && fileInput) {
    uploadZone.onclick = function() { fileInput.click(); };
    fileInput.onchange = function() { if (this.files[0]) handleFileUpload(this.files[0]); };
    uploadZone.ondragover = function(e) { e.preventDefault(); e.stopPropagation(); this.classList.add('dragover'); };
    uploadZone.ondragleave = function(e) { e.preventDefault(); e.stopPropagation(); this.classList.remove('dragover'); };
    uploadZone.ondrop = function(e) {
      e.preventDefault(); e.stopPropagation(); this.classList.remove('dragover');
      if (e.dataTransfer.files[0]) handleFileUpload(e.dataTransfer.files[0]);
    };
  }
  if (resetBtn) {
    resetBtn.onclick = function() {
      if (confirm('Reset to the original embedded data? Your uploaded data will be removed.')) resetData();
    };
  }
})();

// Geo toggle (State / Zone) for Company Deep-Dive
document.getElementById('geoChips-company').querySelectorAll('.view-chip').forEach(chip => {
  chip.onclick = () => {
    currentGeo = chip.dataset.geo;
    document.getElementById('geoChips-company').querySelectorAll('.view-chip').forEach(c => c.classList.toggle('active', c.dataset.geo === currentGeo));
    if (currentTab === 'company') renderCurrentTab();
  };
});

// ============================================
// CHAT WITH DATA
// ============================================
function populateModelDropdown(selEl) {
  selEl.textContent = '';
  Object.keys(CHAT_MODELS).forEach(function(k) {
    var opt = document.createElement('option');
    opt.value = k; opt.textContent = CHAT_MODELS[k].name;
    if (k === selectedModel) opt.selected = true;
    selEl.appendChild(opt);
  });
}

function switchModel(modelKey) {
  if (!CHAT_MODELS[modelKey]) return;
  selectedModel = modelKey;
  try { localStorage.setItem('janchor_selected_model', modelKey); } catch(e) {}
  if (!getChatApiKey()) {
    renderChatTab();
  } else {
    var setupSel = document.getElementById('sel-chat-model');
    var toolbarSel = document.getElementById('sel-toolbar-model');
    if (setupSel) setupSel.value = modelKey;
    if (toolbarSel) toolbarSel.value = modelKey;
  }
}

function updateSetupForModel() {
  var cfg = CHAT_MODELS[selectedModel];
  var inp = document.getElementById('api-key-input');
  var helpLink = document.getElementById('chat-key-link');
  if (inp) inp.placeholder = cfg.keyPlaceholder;
  if (helpLink) { helpLink.href = cfg.keyLink; helpLink.textContent = cfg.keyLinkLabel; }
}

function renderChatTab() {
  const setup = document.getElementById('chat-setup');
  const iface = document.getElementById('chat-interface');
  var setupSel = document.getElementById('sel-chat-model');
  if (setupSel) {
    populateModelDropdown(setupSel);
    setupSel.onchange = function() { selectedModel = this.value; try { localStorage.setItem('janchor_selected_model', this.value); } catch(e) {} updateSetupForModel(); };
  }
  updateSetupForModel();
  if (!getChatApiKey()) { setup.style.display='block'; iface.style.display='none'; return; }
  setup.style.display='none'; iface.style.display='flex';
  var toolbarSel = document.getElementById('sel-toolbar-model');
  if (toolbarSel) {
    populateModelDropdown(toolbarSel);
    toolbarSel.onchange = function() { switchModel(this.value); };
  }
  const msgs = document.getElementById('chat-messages');
  const sugs = document.getElementById('chat-suggestions');
  if (chatHistory.length === 0) {
    sugs.style.display='grid'; while(msgs.firstChild) msgs.removeChild(msgs.firstChild);
  } else {
    sugs.style.display='none';
    renderAllChatMessages();
  }
}

function renderAllChatMessages() {
  const container = document.getElementById('chat-messages');
  container.innerHTML = '';
  chatHistory.forEach(msg => {
    const div = createMsgDiv(msg);
    container.appendChild(div);
  });
  container.scrollTop = container.scrollHeight;
}

function createMsgDiv(msg) {
  const div = document.createElement('div');
  div.className = 'chat-msg ' + msg.role;
  div.id = 'msg-' + msg.id;
  if (msg.role === 'user') {
    div.textContent = msg.content;
  } else {
    const segments = parseAssistantResponse(msg.content);
    const textDiv = document.createElement('div');
    textDiv.className = 'msg-text';
    segments.forEach(seg => renderSegment(seg, textDiv, msg.id));
    div.appendChild(textDiv);
    // Actions
    const actions = document.createElement('div');
    actions.className = 'msg-actions';
    const saveBtn = document.createElement('button');
    saveBtn.className = 'msg-action-btn' + (msg.saved ? ' saved' : '');
    saveBtn.innerHTML = msg.saved ? '&#11088;' : '&#9734;';
    saveBtn.title = msg.saved ? 'Unsave' : 'Save';
    saveBtn.onclick = function(e) { e.stopPropagation(); toggleSaveMessage(msg.id); };
    actions.appendChild(saveBtn);
    div.appendChild(actions);
  }
  return div;
}

function simpleMarkdown(text) {
  let html = text
    .replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;')
    .replace(/\\*\\*(.+?)\\*\\*/g,'<b>$1</b>')
    .replace(/\\*(.+?)\\*/g,'<i>$1</i>')
    .replace(/`([^`]+)`/g,'<code>$1</code>')
    .replace(/^[\\-\\*] (.+)$/gm,'<li>$1</li>')
    .replace(/^(\\d+)\\. (.+)$/gm,'<li>$2</li>');
  // Wrap consecutive <li> in <ul>
  html = html.replace(/((?:<li>.*?<\\/li>\\s*)+)/g,'<ul>$1</ul>');
  html = html.replace(/\\n\\n/g,'</p><p>').replace(/\\n/g,'<br>');
  return '<p>' + html + '</p>';
}

function parseAssistantResponse(text) {
  const segments = [];
  const regex = /```(js|javascript|chart|table)\\n([\\s\\S]*?)```/g;
  let last = 0, match;
  while ((match = regex.exec(text)) !== null) {
    if (match.index > last) segments.push({type:'text', content:text.slice(last, match.index)});
    let type = match[1];
    if (type === 'javascript') type = 'js';
    segments.push({type, content:match[2]});
    last = regex.lastIndex;
  }
  if (last < text.length) segments.push({type:'text', content:text.slice(last)});
  return segments;
}

function renderSegment(seg, container, msgId) {
  if (seg.type === 'text') {
    const d = document.createElement('div');
    d.innerHTML = simpleMarkdown(seg.content.trim());
    container.appendChild(d);
  } else if (seg.type === 'js') {
    const result = executeChatCode(seg.content);
    const d = document.createElement('div');
    d.className = 'chat-code-result' + (result.success ? '' : ' chat-code-error');
    if (result.success) {
      d.textContent = result.result !== undefined ? (typeof result.result === 'object' ? JSON.stringify(result.result, null, 2) : String(result.result)) : '(executed)';
    } else {
      d.textContent = 'Error: ' + result.error;
    }
    container.appendChild(d);
  } else if (seg.type === 'chart') {
    const chartDiv = document.createElement('div');
    chartDiv.className = 'chat-chart-container';
    const chartId = 'chat-chart-' + msgId + '-' + Date.now() + '-' + Math.random().toString(36).substr(2,4);
    const plotDiv = document.createElement('div');
    plotDiv.id = chartId;
    plotDiv.style.height = '400px';
    chartDiv.appendChild(plotDiv);
    // Copy chart button
    const copyBtn = document.createElement('button');
    copyBtn.className = 'chat-chart-copy';
    copyBtn.textContent = '📷 Copy';
    copyBtn.title = 'Copy chart as image';
    copyBtn.onclick = function(e) {
      e.stopPropagation();
      Plotly.toImage(chartId, {format:'png', width:900, height:500}).then(function(url){
        fetch(url).then(r=>r.blob()).then(function(blob){
          try { navigator.clipboard.write([new ClipboardItem({'image/png':blob})]); copyBtn.textContent='✓ Copied'; setTimeout(function(){copyBtn.textContent='📷 Copy';},1500); }
          catch(err){ var w=window.open(''); w.document.write('<img src="'+url+'">'); }
        });
      });
    };
    chartDiv.appendChild(copyBtn);
    container.appendChild(chartDiv);
    try {
      const spec = JSON.parse(seg.content);
      const traces = spec.data || spec.traces || [];
      const layout = {...PLOTLY_LAYOUT, margin:{l:70,r:30,t:30,b:90}, legend:{orientation:'h',y:-0.18,x:0.5,xanchor:'center',font:{size:10}}, ...(spec.layout||{})};
      setTimeout(function(){ Plotly.newPlot(chartId, traces, layout, PLOTLY_CONFIG); }, 50);
    } catch(e) {
      plotDiv.textContent = 'Chart error: ' + e.message;
      plotDiv.style.color = '#dc2626';
      plotDiv.style.padding = '20px';
    }
  } else if (seg.type === 'table') {
    const d = document.createElement('div');
    d.className = 'chat-table-container';
    try {
      const spec = JSON.parse(seg.content);
      let html = '<table><thead><tr>' + (spec.headers||[]).map(h=>'<th>'+h+'</th>').join('') + '</tr></thead><tbody>';
      (spec.rows||[]).forEach(row => {
        html += '<tr>' + row.map(c=>'<td'+(typeof c==='number'?' class="align-right"':'')+'>'+c+'</td>').join('') + '</tr>';
      });
      html += '</tbody></table>';
      d.innerHTML = html;
    } catch(e) {
      d.textContent = 'Table error: ' + e.message;
    }
    container.appendChild(d);
  }
}

function executeChatCode(code) {
  try {
    var result = eval(code);
    return {success:true, result:result};
  } catch(e) {
    return {success:false, error:e.message};
  }
}

function buildDataSummary() {
  var lines = [];
  lines.push('Current segment: ' + currentSegment);
  lines.push('Subsegments: ' + segSubsegs.join(', '));
  lines.push('Companies (' + segCompanies.length + '): ' + segCompanies.join(', '));
  lines.push('States (' + segStates.length + '): ' + segStates.join(', '));
  lines.push('Zones (' + segZones.length + '): ' + segZones.join(', '));
  lines.push('Zone-State mapping: ' + segZones.map(function(z){ return z + ': ' + (zoneStateMap[z]||[]).join(', '); }).join(' | '));
  return lines.join('\\n');
}

function buildFullSegmentData() {
  var lines = [];
  // Header with all quarter labels
  lines.push('Subsegment|Zone|State|Manufacturer|' + QLABELS.join('|'));
  // Industry totals row (sum of all companies) for easy market share computation
  var indVols = getIndustryVols('All');
  var indRow = ['ALL|ALL|ALL|INDUSTRY TOTAL'];
  for (var q = 0; q < NQ; q++) indRow.push(Math.round(indVols[q]));
  lines.push(indRow.join('|'));
  // All data rows for current segment
  for (var i = 0; i < ROWS.length; i++) {
    if (ROWS[i][0] !== currentSegment) continue;
    var r = ROWS[i];
    var row = [r[1], r[2], r[3], r[4]];
    for (var q = 0; q < NQ; q++) row.push(r[5 + q]);
    lines.push(row.join('|'));
  }
  return lines.join('\\n');
}

function buildCrossSegSummary() {
  var allSegs = ['PV','2W','MHCV','LCV','3W'];
  var otherSegs = allSegs.filter(function(s) { return s !== currentSegment && segCompanies.length > 0; });
  // Only include segments that exist in the data
  otherSegs = otherSegs.filter(function(seg) {
    for (var i = 0; i < ROWS.length; i++) { if (ROWS[i][0] === seg) return true; }
    return false;
  });
  if (otherSegs.length === 0) return '';
  var fyHeaders = FYS.map(function(fy) {
    return FY_Q_IDXS[fy].length < 4 ? fy + '(' + FY_Q_IDXS[fy].length + 'Q)' : fy;
  });
  var lines = ['=== OTHER SEGMENTS (industry annual totals for cross-reference) ==='];
  lines.push('Segment|' + fyHeaders.join('|'));
  otherSegs.forEach(function(seg) {
    var sv = new Array(NQ).fill(0);
    for (var i = 0; i < ROWS.length; i++) {
      if (ROWS[i][0] === seg) {
        for (var q = 0; q < NQ; q++) sv[q] += ROWS[i][5 + q];
      }
    }
    var annSv = FYS.map(function(fy) {
      return Math.round(FY_Q_IDXS[fy].reduce(function(s, qi) { return s + sv[qi]; }, 0));
    });
    lines.push(seg + '|' + annSv.join('|'));
  });
  return lines.join('\\n');
}

function buildSystemPrompt() {
  var nSegRows = 0;
  for (var i = 0; i < ROWS.length; i++) {
    if (ROWS[i][0] === currentSegment) nSegRows++;
  }
  return 'You are an expert analyst for Indian auto industry state-wise primary sales data.\\n\\n' +
  'IMPORTANT: You have the COMPLETE raw dataset for the ' + currentSegment + ' segment below (' + nSegRows + ' rows x ' + NQ + ' quarters). ' +
  'This is ALL the data — every company, every state, every quarter. Read actual numbers from this table to answer questions. ' +
  'Only use ```js blocks for complex computations (CAGR, custom aggregations).\\n\\n' +
  'DATA OVERVIEW:\\n' + buildDataSummary() + '\\n\\n' +
  'FISCAL YEAR REFERENCE:\\n' +
  '- Indian FY: FY17 = Apr 2016 to Mar 2017. Quarters: Q1=Apr-Jun, Q2=Jul-Sep, Q3=Oct-Dec, Q4=Jan-Mar\\n' +
  '- FYS: ' + JSON.stringify(FYS) + '\\n' +
  '- ' + FYS[NFY-1] + ' is PARTIAL (' + (FY_Q_IDXS[FYS[NFY-1]]||[]).length + ' quarters only)\\n' +
  '- To compute annual FY total: sum all quarters in that FY (e.g. FY20 = Q1FY20+Q2FY20+Q3FY20+Q4FY20)\\n' +
  '- Values are unit volumes (individual vehicles, not in lakhs/crores)\\n' +
  '- Data is pipe-delimited (|)\\n\\n' +
  '=== COMPLETE RAW DATA: ' + currentSegment + ' SEGMENT (' + nSegRows + ' rows) ===\\n' +
  'Each row = one unique (Subsegment, Zone, State, Manufacturer) combination with quarterly volumes.\\n' +
  'First data row is INDUSTRY TOTAL (sum of all companies) for market share computation.\\n' +
  buildFullSegmentData() + '\\n\\n' +
  buildCrossSegSummary() + '\\n\\n' +
  'JS HELPER FUNCTIONS (for complex queries only, use in ```js blocks):\\n' +
  '- getIndustryVols(sub), getCompanyVols(co, sub) -> quarterly arrays\\n' +
  '- getStateIndustryVols(state, sub), getStateCompanyVols(state, co, sub)\\n' +
  '- getZoneIndustryVols(zone, sub), getZoneCompanyVols(zone, co, sub)\\n' +
  '- filterRows(company, state, subseg, zone) -> row indices\\n' +
  '- sumVolumes(rowIdxs) -> quarterly array. annualVols(qVols) -> annual FY array\\n' +
  '- computeShare(coVols, indVols) -> percentage array. fmt(n) -> formatted string\\n' +
  '- NQ=' + NQ + ', NFY=' + NFY + ', QLABELS, PALETTE, COMPANY_COLORS\\n\\n' +
  'OUTPUT FORMATS:\\n' +
  '1. Text: use **bold** and *italic*. Cite numbers from the data.\\n' +
  '2. ```js block: JavaScript executed in browser. Last expression shown as result.\\n' +
  '3. ```chart block: Plotly JSON {"data":[traces],"layout":{}}.\\n' +
  '4. ```table block: {"headers":["A","B"],"rows":[[1,2],[3,4]]}\\n\\n' +
  'RULES:\\n' +
  '- ALWAYS read numbers from the raw data table. You have ALL ' + nSegRows + ' rows for ' + currentSegment + '.\\n' +
  '- For market share: company_vol / INDUSTRY_TOTAL * 100\\n' +
  '- For YoY growth: (current - previous) / previous * 100\\n' +
  '- For annual data: sum the quarterly columns within each FY\\n' +
  '- For state/zone/subsegment analysis: filter rows by the relevant column\\n' +
  '- For charts, use COMPANY_COLORS[name] or PALETTE[i]\\n' +
  '- Keep responses concise and data-driven. Always cite specific numbers.\\n' +
  '- For FYTD comparisons with partial years, compare same quarters only.';
}

function addChatMessage(role, content, saved) {
  var msg = {
    id: 'msg_' + Date.now() + '_' + Math.random().toString(36).substr(2,5),
    role: role,
    content: content,
    timestamp: new Date().toISOString(),
    saved: saved || false
  };
  chatHistory.push(msg);
  saveChatHistory();
  return msg;
}

function saveChatHistory() {
  try { localStorage.setItem('janchor_chat_history', JSON.stringify(chatHistory)); } catch(e) {}
}

function showTypingIndicator() {
  var container = document.getElementById('chat-messages');
  var div = document.createElement('div');
  div.className = 'typing-indicator';
  div.id = 'typing-indicator';
  div.innerHTML = '<div class="typing-dot"></div><div class="typing-dot"></div><div class="typing-dot"></div>';
  container.appendChild(div);
  container.scrollTop = container.scrollHeight;
}

function removeTypingIndicator() {
  var el = document.getElementById('typing-indicator');
  if (el) el.remove();
}

function buildApiRequest(systemPrompt, messages) {
  var cfg = CHAT_MODELS[selectedModel];
  var key = getChatApiKey();
  if (cfg.provider === 'anthropic') {
    return {
      url: 'https://api.anthropic.com/v1/messages',
      headers: {
        'Content-Type': 'application/json',
        'x-api-key': key,
        'anthropic-version': '2023-06-01',
        'anthropic-dangerous-direct-browser-access': 'true'
      },
      body: {
        model: cfg.model,
        max_tokens: 4096,
        system: [{type:'text', text: systemPrompt, cache_control:{type:'ephemeral'}}],
        messages: messages
      }
    };
  } else if (cfg.provider === 'google') {
    var contents = messages.map(function(m) {
      return { role: m.role === 'assistant' ? 'model' : 'user', parts: [{text: m.content}] };
    });
    return {
      url: 'https://generativelanguage.googleapis.com/v1beta/models/' + cfg.model + ':generateContent?key=' + key,
      headers: { 'Content-Type': 'application/json' },
      body: {
        systemInstruction: { parts: [{text: systemPrompt}] },
        contents: contents,
        generationConfig: { maxOutputTokens: 8192 }
      }
    };
  }
}

function parseApiResponse(provider, data) {
  if (provider === 'anthropic') {
    return (data.content && data.content[0] && data.content[0].text) || 'No response received.';
  } else if (provider === 'google') {
    return (data.candidates && data.candidates[0] && data.candidates[0].content &&
      data.candidates[0].content.parts && data.candidates[0].content.parts[0] &&
      data.candidates[0].content.parts[0].text) || 'No response received.';
  }
  return 'No response received.';
}

function logApiUsage(provider, data) {
  if (provider === 'anthropic' && data.usage) {
    var u = data.usage;
    var cacheRead = u.cache_read_input_tokens || 0;
    var cacheCreate = u.cache_creation_input_tokens || 0;
    var totalIn = u.input_tokens || 0;
    if (cacheRead > 0) {
      console.log('[Chat Cost] ' + CHAT_MODELS[selectedModel].name + ' | Input: ' + totalIn + ' tokens, Cache READ: ' + cacheRead + ' (90% cheaper), Output: ' + (u.output_tokens||0));
    } else if (cacheCreate > 0) {
      console.log('[Chat Cost] ' + CHAT_MODELS[selectedModel].name + ' | Input: ' + totalIn + ' tokens, Cache WRITE: ' + cacheCreate + ' (first call), Output: ' + (u.output_tokens||0));
    } else {
      console.log('[Chat Cost] ' + CHAT_MODELS[selectedModel].name + ' | Input: ' + totalIn + ' tokens, Output: ' + (u.output_tokens||0));
    }
  } else if (provider === 'google' && data.usageMetadata) {
    var g = data.usageMetadata;
    console.log('[Chat Cost] ' + CHAT_MODELS[selectedModel].name + ' | Input: ' + (g.promptTokenCount||0) + ' tokens, Output: ' + (g.candidatesTokenCount||0) + ' tokens, Total: ' + (g.totalTokenCount||0));
  }
}

async function sendChatMessage(userText) {
  if (!userText || !userText.trim()) return;
  userText = userText.trim();
  document.getElementById('chat-suggestions').style.display = 'none';

  // Add user message
  var userMsg = addChatMessage('user', userText);
  var container = document.getElementById('chat-messages');
  container.appendChild(createMsgDiv(userMsg));
  container.scrollTop = container.scrollHeight;

  // Clear input
  document.getElementById('chat-input').value = '';
  document.getElementById('chat-input').style.height = 'auto';

  showTypingIndicator();

  // Build messages for API (last 20 messages)
  var apiMessages = chatHistory.slice(-21, -1).concat([{role:'user',content:userText}]).map(function(m){
    return {role: m.role, content: m.content};
  }).filter(function(m){ return m.role === 'user' || m.role === 'assistant'; });
  // Ensure alternating user/assistant
  var cleaned = [];
  for (var i = 0; i < apiMessages.length; i++) {
    if (cleaned.length === 0 || cleaned[cleaned.length-1].role !== apiMessages[i].role) {
      cleaned.push(apiMessages[i]);
    }
  }
  if (cleaned.length > 0 && cleaned[0].role !== 'user') cleaned.shift();

  try {
    var cfg = CHAT_MODELS[selectedModel];
    var apiReq = buildApiRequest(buildSystemPrompt(), cleaned);
    var response = await fetch(apiReq.url, {
      method: 'POST',
      headers: apiReq.headers,
      body: JSON.stringify(apiReq.body)
    });

    removeTypingIndicator();

    if (!response.ok) {
      var errData = {};
      try { errData = await response.json(); } catch(e2) {}
      var errMsg = response.status === 401 || response.status === 403 ? 'Invalid API key. Please check your ' + cfg.name + ' key.'
        : response.status === 429 ? 'Rate limited. Please wait a moment and try again.'
        : response.status === 529 ? cfg.name + ' is overloaded. Please try again shortly.'
        : 'API error (' + response.status + '): ' + (errData.error?.message || (errData.error?.errors && errData.error.errors[0]?.message) || 'Unknown error');
      var errMsgObj = addChatMessage('assistant', errMsg);
      container.appendChild(createMsgDiv(errMsgObj));
      container.scrollTop = container.scrollHeight;
      return;
    }

    var data = await response.json();
    var assistantText = parseApiResponse(cfg.provider, data);
    logApiUsage(cfg.provider, data);

    var assistantMsg = addChatMessage('assistant', assistantText);
    container.appendChild(createMsgDiv(assistantMsg));
    container.scrollTop = container.scrollHeight;

  } catch(err) {
    removeTypingIndicator();
    var networkErr = addChatMessage('assistant', 'Network error: ' + err.message + '. Check your internet connection.');
    container.appendChild(createMsgDiv(networkErr));
    container.scrollTop = container.scrollHeight;
  }
}

function toggleSaveMessage(msgId) {
  var newSaved = false;
  for (var i = 0; i < chatHistory.length; i++) {
    if (chatHistory[i].id === msgId) {
      chatHistory[i].saved = !chatHistory[i].saved;
      newSaved = chatHistory[i].saved;
      break;
    }
  }
  saveChatHistory();
  // Update star in-place without re-rendering (avoids scroll jump)
  var msgDiv = document.getElementById('msg-' + msgId);
  if (msgDiv) {
    var btn = msgDiv.querySelector('.msg-action-btn');
    if (btn) {
      btn.className = 'msg-action-btn' + (newSaved ? ' saved' : '');
      btn.innerHTML = newSaved ? '&#11088;' : '&#9734;';
      btn.title = newSaved ? 'Unsave' : 'Save';
    }
  }
}

function exportSavedMessages() {
  var saved = chatHistory.filter(function(m){ return m.saved; });
  if (saved.length === 0) { alert('No saved messages to export. Click the star icon on messages to save them.'); return; }
  // Collect chart image promises
  var chartImages = {};
  var chartPromises = [];
  // Find all rendered chart divs from saved messages
  saved.forEach(function(m) {
    if (m.role !== 'assistant') return;
    var msgDiv = document.getElementById('msg-' + m.id);
    if (!msgDiv) return;
    var charts = msgDiv.querySelectorAll('[id^="chat-chart-"]');
    charts.forEach(function(el, ci) {
      var pid = el.id;
      if (el.querySelector('.js-plotly-plot')) {
        chartPromises.push(
          Plotly.toImage(pid, {format:'png', width:900, height:500}).then(function(dataUrl) {
            if (!chartImages[m.id]) chartImages[m.id] = [];
            chartImages[m.id].push(dataUrl);
          }).catch(function(){ })
        );
      }
    });
  });
  Promise.all(chartPromises).then(function() {
    var parts = [];
    parts.push('<html><head><meta charset="utf-8"><title>Auto Insights - ' + currentSegment + '</title>');
    parts.push('<style>body{font-family:system-ui,-apple-system,sans-serif;max-width:900px;margin:40px auto;padding:0 20px;color:#1f2937;line-height:1.6}');
    parts.push('h1{color:#1e40af;border-bottom:2px solid #2563eb;padding-bottom:8px}h2{color:#374151;margin-top:32px}');
    parts.push('.q{background:#eff6ff;border-left:4px solid #2563eb;padding:12px 16px;margin:16px 0;border-radius:0 8px 8px 0;font-weight:500}');
    parts.push('table{border-collapse:collapse;width:100%;margin:12px 0;font-size:13px}th,td{border:1px solid #d1d5db;padding:6px 10px;text-align:left}th{background:#f1f5f9;font-weight:600}');
    parts.push('img{max-width:100%;border-radius:8px;margin:12px 0;box-shadow:0 2px 8px rgba(0,0,0,0.1)}');
    parts.push('hr{border:none;border-top:1px solid #e5e7eb;margin:24px 0}.meta{color:#6b7280;font-size:12px}</style></head><body>');
    parts.push('<h1>Saved Chat Insights - ' + currentSegment + ' Segment</h1>');
    parts.push('<p class="meta">Exported: ' + new Date().toLocaleString() + '</p>');
    var insightNum = 0;
    saved.forEach(function(m) {
      if (m.role === 'user') {
        parts.push('<div class="q">' + m.content.replace(/</g,'&lt;').replace(/>/g,'&gt;') + '</div>');
      } else {
        insightNum++;
        parts.push('<h2>Insight ' + insightNum + '</h2>');
        var segments = parseAssistantResponse(m.content);
        var chartIdx = 0;
        segments.forEach(function(seg) {
          if (seg.type === 'text') {
            parts.push(simpleMarkdown(seg.content.trim()));
          } else if (seg.type === 'table') {
            try {
              var spec = JSON.parse(seg.content);
              var html = '<table><thead><tr>' + (spec.headers||[]).map(function(h){return '<th>'+h+'</th>';}).join('') + '</tr></thead><tbody>';
              (spec.rows||[]).forEach(function(row) {
                html += '<tr>' + row.map(function(c){return '<td>'+c+'</td>';}).join('') + '</tr>';
              });
              html += '</tbody></table>';
              parts.push(html);
            } catch(e) {}
          } else if (seg.type === 'chart') {
            if (chartImages[m.id] && chartImages[m.id][chartIdx]) {
              parts.push('<img src="' + chartImages[m.id][chartIdx] + '" alt="Chart">');
            } else {
              parts.push('<p><em>[Chart - view in dashboard]</em></p>');
            }
            chartIdx++;
          }
          // js segments are intentionally skipped
        });
        parts.push('<hr>');
      }
    });
    parts.push('</body></html>');
    var blob = new Blob([parts.join('\\n')], {type:'text/html'});
    var a = document.createElement('a');
    a.href = URL.createObjectURL(blob);
    a.download = 'auto_insights_' + currentSegment + '_' + new Date().toISOString().slice(0,10) + '.html';
    a.click();
    URL.revokeObjectURL(a.href);
  });
}

function clearChat() {
  if (chatHistory.length === 0) return;
  if (!confirm('Clear all chat messages? Saved messages will also be removed.')) return;
  // Purge any Plotly charts
  document.querySelectorAll('[id^="chat-chart-"]').forEach(function(el){ try{Plotly.purge(el.id);}catch(e){} });
  chatHistory = [];
  saveChatHistory();
  renderChatTab();
}

function removeChatApiKey() {
  var cfg = CHAT_MODELS[selectedModel];
  if (!confirm('Remove your ' + cfg.name + ' API key? You will need to re-enter it to use chat with this provider.')) return;
  delete chatApiKeys[cfg.provider];
  try { localStorage.setItem('janchor_api_keys', JSON.stringify(chatApiKeys)); } catch(e) {}
  renderChatTab();
}

// Chat event listeners
(function() {
  document.getElementById('btn-save-key').onclick = function() {
    var key = document.getElementById('api-key-input').value.trim();
    var cfg = CHAT_MODELS[selectedModel];
    if (!key || !key.startsWith(cfg.keyPrefix)) {
      alert('Please enter a valid ' + cfg.name + ' API key (starts with ' + cfg.keyPrefix + '...)');
      return;
    }
    chatApiKeys[cfg.provider] = key;
    try { localStorage.setItem('janchor_api_keys', JSON.stringify(chatApiKeys)); } catch(e) {}
    renderChatTab();
  };
  document.getElementById('btn-chat-send').onclick = function() {
    sendChatMessage(document.getElementById('chat-input').value);
  };
  document.getElementById('chat-input').addEventListener('keydown', function(e) {
    if (e.key === 'Enter' && !e.shiftKey) {
      e.preventDefault();
      sendChatMessage(this.value);
    }
  });
  // Auto-resize textarea
  document.getElementById('chat-input').addEventListener('input', function() {
    this.style.height = 'auto';
    this.style.height = Math.min(this.scrollHeight, 120) + 'px';
  });
  document.getElementById('btn-chat-clear').onclick = clearChat;
  document.getElementById('btn-chat-export').onclick = exportSavedMessages;
  document.getElementById('btn-chat-remove-key').onclick = removeChatApiKey;
  document.getElementById('btn-chat-download-data').onclick = function() {
    var csv = 'Segment,Subsegment,Zone,State,Manufacturer,' + QLABELS.join(',') + '\\n';
    ROWS.forEach(function(r) {
      var meta = r.slice(0,5).map(function(v){ return '"' + v + '"'; });
      csv += meta.join(',') + ',' + r.slice(5).join(',') + '\\n';
    });
    var blob = new Blob([csv], {type:'text/csv;charset=utf-8'});
    var a = document.createElement('a');
    a.href = URL.createObjectURL(blob);
    a.download = 'janchor_auto_data_all_segments.csv';
    document.body.appendChild(a); a.click(); document.body.removeChild(a);
  };
  // Suggestion cards
  document.querySelectorAll('.chat-suggestion').forEach(function(card) {
    card.onclick = function() {
      var prompt = this.dataset.prompt;
      document.getElementById('chat-input').value = prompt;
      sendChatMessage(prompt);
    };
  });
})();

// ============================================
// INITIALIZE
// ============================================
buildIndexes();
updateDropdowns();
['overview','company','state','zone'].forEach(tab => populatePeriodSelector(tab));
renderSubsegChips();
syncViewChips();
renderCurrentTab();
</script>
</body>
</html>'''

with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
    f.write(html)

print(f"Dashboard generated: {OUTPUT_FILE}")
print(f"File size: {os.path.getsize(OUTPUT_FILE)/1024:.0f} KB")
